import io
import pandas as pd
import logging
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Any, Tuple
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Set up logger
logger = logging.getLogger(__name__)

def export_to_excel(df: pd.DataFrame, sheet_name: str = "Data") -> Tuple[bytes, List[str]]:
    """
    Export a DataFrame to an Excel file.
    
    Args:
        df: DataFrame to export
        sheet_name: Name of the sheet in the Excel file
        
    Returns:
        Excel file as bytes
    """
    # Create a buffer to store the Excel file
    buffer = io.BytesIO()
    
    # Create Excel writer
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Write DataFrame to Excel
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Auto-adjust columns' width
        worksheet = writer.sheets[sheet_name]
        for i, col in enumerate(df.columns):
            max_length = max(
                df[col].astype(str).apply(len).max(),
                len(str(col))
            ) + 2
            worksheet.column_dimensions[get_column_letter(i+1)].width = max_length
    
    # Seek to the beginning of the buffer
    buffer.seek(0)
    
    # For simple generic exports, we return an empty alerts list for consistency
    alerts = []
    return buffer.getvalue(), alerts

def create_monthly_financial_report(invoice_data: list, month: str = None, year: str = None) -> Tuple[bytes, List[str]]:
    """
    Create a monthly financial report in Excel format with income and expenses,
    matching the exact format from the provided example.
    
    Args:
        invoice_data: List of invoice data dictionaries
        month: Month for the report (default: current month)
        year: Year for the report (default: current year)
        
    Returns:
        Excel file as bytes
    """
    # Set default month and year if not provided
    if month is None:
        month = datetime.now().strftime("%B")
    if year is None:
        year = str(datetime.now().year)
    
    # Create a buffer to store the Excel file
    buffer = io.BytesIO()
    
    # Create income and expense data
    income_data = []
    expense_data = []
    
    for invoice in invoice_data:
        # Get transaction type - check both standard and raw extraction fields
        transaction_type = invoice.get('transaction_type', '') or invoice.get('Kind of Transaction', '')
        if not transaction_type and invoice.get('raw_extraction'):
            transaction_type = invoice['raw_extraction'].get('Kind of Transaction', '')
        
        transaction_type = str(transaction_type).upper() if transaction_type else ''
        
        # Check if this is an income entry
        if 'INCOME' in transaction_type or 'COLLECTION' in transaction_type or 'SALES' in transaction_type:
            # For income, format according to INCOME AND COLLECTIONS sheet
            income_item = {}
            
            # First try to get data from raw extraction if available
            if invoice.get('raw_extraction'):
                raw = invoice['raw_extraction']
                income_item = {
                    "DATE": raw.get('Date'),
                    "KIND OF TRANSACTION": raw.get('Kind of Transaction'),
                    "TRANSACTORS": raw.get('Transactor'),
                    "DESCRIPTION": raw.get('Description'),
                    "KIND OF DOCUMENT": "Invoice",  # Default value if not present
                    "NUMBER OF INVOICE": raw.get('Number of Invoice'),
                    "NET AMOUNT": raw.get('Net'),
                    "VAT": raw.get('VAT'),
                    "TOTAL AMOUNT": raw.get('Total BGN') or raw.get('Total Euro'),
                    "PAYMENT": raw.get('Payment Mode'),
                    "INFORMATION": raw.get('Notes or Payment Date')
                }
            else:
                # Fall back to the standard structure if no raw extraction
                # Handle category (use categories list if available)
                category = ''
                if 'categories' in invoice and invoice['categories']:
                    category = ', '.join(invoice['categories'])
                elif 'category' in invoice and invoice['category']:
                    category = invoice['category']
                
                income_item = {
                    "DATE": invoice.get('invoice_date'),
                    "KIND OF TRANSACTION": transaction_type,
                    "TRANSACTORS": invoice.get('supplier_name'),
                    "DESCRIPTION": invoice.get('description'),
                    "KIND OF DOCUMENT": "Invoice",  # Default value
                    "NUMBER OF INVOICE": invoice.get('invoice_number'),
                    "NET AMOUNT": invoice.get('amount'),
                    "VAT": invoice.get('tax_amount'),
                    "TOTAL AMOUNT": invoice.get('total_amount'),
                    "PAYMENT": invoice.get('payment_method'),
                    "INFORMATION": invoice.get('notes')
                }
            
            income_data.append(income_item)
            
        else:
            # For expenses, format according to EXPENSES AND PAYMENTS sheet with only the required fields
            expense_item = {}
            
            # First try to get data from raw extraction if available
            if invoice.get('raw_extraction'):
                raw = invoice['raw_extraction']
                # Get transaction type
                transaction_type = raw.get('Kind of Transaction', '')
                
                # Get supplier name/transactor
                transactor = raw.get('Transactor', '')
                
                # Get expense category
                category = raw.get('Expense Category Account', '')
                
                # Get VAT number instead of VAT amount
                vat_number = raw.get('VAT Number', '') or raw.get('VAT', '')
                
                expense_item = {
                    "Kind of Transaction": transaction_type,
                    "Transactor": transactor,
                    "Expense Category Account": category,
                    "VAT Number": vat_number
                }
            else:
                # Fall back to the standard structure if no raw extraction
                # Handle category (use categories list if available)
                category = ''
                if 'categories' in invoice and invoice['categories']:
                    category = ', '.join(invoice['categories'])
                elif 'category' in invoice and invoice['category']:
                    category = invoice['category']
                
                # Get VAT number instead of tax amount
                vat_number = invoice.get('vat_number') or invoice.get('tax_amount', '')
                
                expense_item = {
                    "Kind of Transaction": transaction_type,
                    "Transactor": invoice.get('supplier_name', ''),
                    "Expense Category Account": category,
                    "VAT Number": vat_number
                }
            
            expense_data.append(expense_item)
    
    # Income columns from the example
    income_columns = [
        "DATE", "KIND OF TRANSACTION", "TRANSACTORS", "DESCRIPTION", 
        "KIND OF DOCUMENT", "NUMBER OF INVOICE", "NET AMOUNT", 
        "VAT", "TOTAL AMOUNT", "PAYMENT", "INFORMATION"
    ]
    
    # Simplified expense columns focusing on the four key fields
    expense_columns = [
        "Kind of Transaction", "Transactor", "Expense Category Account", "VAT Number"
    ]
    
    # Convert to DataFrames
    income_df = pd.DataFrame(income_data, columns=income_columns) if income_data else pd.DataFrame(columns=income_columns)
    expense_df = pd.DataFrame(expense_data, columns=expense_columns) if expense_data else pd.DataFrame(columns=expense_columns)
    
    # Create Excel writer
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Write title and headers with formatting
        workbook = writer.book
        
        # Create INCOME AND COLLECTIONS sheet
        income_sheet_name = "INCOME AND COLLECTIONS"
        if not income_df.empty:
            income_df.to_excel(writer, sheet_name=income_sheet_name, index=False)
        else:
            pd.DataFrame(columns=income_columns).to_excel(
                writer, sheet_name=income_sheet_name, index=False)
        
        # Format INCOME sheet
        income_sheet = writer.sheets[income_sheet_name]
        
        # Apply formatting to header row
        header_fill = PatternFill(start_color="E6F2F5", end_color="E6F2F5", fill_type="solid")
        for col_num in range(1, len(income_columns) + 1):
            cell = income_sheet.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        
        # Add borders to all cells
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
                             
        # Set column widths and apply borders
        for i, col in enumerate(income_columns, 1):
            column_letter = get_column_letter(i)
            # Set appropriate column width based on content
            if col in ['DATE', 'KIND OF TRANSACTION', 'PAYMENT']:
                income_sheet.column_dimensions[column_letter].width = 15
            elif col in ['TRANSACTORS', 'DESCRIPTION']:
                income_sheet.column_dimensions[column_letter].width = 25
            elif col in ['NUMBER OF INVOICE', 'INFORMATION', 'KIND OF DOCUMENT']:
                income_sheet.column_dimensions[column_letter].width = 20
            else:  # Numeric columns
                income_sheet.column_dimensions[column_letter].width = 12
                
            # Apply borders and alignment to all cells in the column
            for row in range(1, len(income_data) + 2):  # +2 for header row
                cell = income_sheet.cell(row=row, column=i)
                cell.border = thin_border
                if col in ['NET AMOUNT', 'VAT', 'TOTAL AMOUNT']:
                    cell.alignment = Alignment(horizontal='right')
        
        # Create EXPENSES AND PAYMENTS sheet
        expense_sheet_name = "EXPENSES AND PAYMENTS"
        if not expense_df.empty:
            expense_df.to_excel(writer, sheet_name=expense_sheet_name, index=False)
        else:
            pd.DataFrame(columns=expense_columns).to_excel(
                writer, sheet_name=expense_sheet_name, index=False)
        
        # Format EXPENSES sheet
        expense_sheet = writer.sheets[expense_sheet_name]
        
        # Apply formatting to header row
        for col_num in range(1, len(expense_columns) + 1):
            cell = expense_sheet.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
                
        # Set column widths and apply borders
        for i, col in enumerate(expense_columns, 1):
            column_letter = get_column_letter(i)
            # Set appropriate column width based on content
            if col in ['Date', 'Kind of Transaction', 'currency', 'Payment mode']:
                expense_sheet.column_dimensions[column_letter].width = 15
            elif col in ['Transactor', 'Expense Category Account', 'description']:
                expense_sheet.column_dimensions[column_letter].width = 25
            elif col in ['NUMBER OF INVOICE', 'Notes or Payment Date']:
                expense_sheet.column_dimensions[column_letter].width = 20
            else:  # Numeric columns
                expense_sheet.column_dimensions[column_letter].width = 12
                
            # Apply borders and alignment to all cells in the column
            for row in range(1, len(expense_data) + 2):  # +2 for header row
                cell = expense_sheet.cell(row=row, column=i)
                cell.border = thin_border
                if col in ['net', 'Vat', 'total BGN', 'total Euro']:
                    cell.alignment = Alignment(horizontal='right')
        
        # Create SUMMARY sheet
        summary_sheet_name = "SUMMARY"
        summary_sheet = workbook.create_sheet(summary_sheet_name)
        
        # Format SUMMARY sheet
        summary_sheet.cell(row=1, column=1).value = f"SUMMARY {month} {year}"
        summary_sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
        summary_sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')
        summary_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        
        # Add table headers
        summary_sheet.cell(row=3, column=1).value = "Category"
        summary_sheet.cell(row=3, column=2).value = "BGN"
        summary_sheet.cell(row=3, column=3).value = "Euro"
        summary_sheet.cell(row=3, column=4).value = "Notes"
        
        # Format the headers
        for col in range(1, 5):
            cell = summary_sheet.cell(row=3, column=col)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Set column widths
        summary_sheet.column_dimensions['A'].width = 25
        summary_sheet.column_dimensions['B'].width = 15
        summary_sheet.column_dimensions['C'].width = 15
        summary_sheet.column_dimensions['D'].width = 25
        
        # Calculate total income BGN and Euro
        total_income_bgn = 0
        total_income_euro = 0
        if not income_df.empty:
            # For income we need to parse the TOTAL AMOUNT
            if 'TOTAL AMOUNT' in income_df.columns:
                income_df['TOTAL AMOUNT'] = pd.to_numeric(income_df['TOTAL AMOUNT'], errors='coerce').fillna(0)
                # We need to determine which are BGN and which are EUR
                # For now, let's assume all are BGN
                total_income_bgn = income_df['TOTAL AMOUNT'].sum()
            
        # Calculate total expenses BGN and Euro
        total_expenses_bgn = 0
        total_expenses_euro = 0
        if not expense_df.empty:
            # For expenses we have separate columns for BGN and Euro
            if 'total BGN' in expense_df.columns:
                expense_df['total BGN'] = pd.to_numeric(expense_df['total BGN'], errors='coerce').fillna(0)
                total_expenses_bgn = expense_df['total BGN'].sum()
            
            if 'total Euro' in expense_df.columns:
                expense_df['total Euro'] = pd.to_numeric(expense_df['total Euro'], errors='coerce').fillna(0)
                total_expenses_euro = expense_df['total Euro'].sum()
            
        # Add income row
        summary_sheet.cell(row=4, column=1).value = "INCOME"
        summary_sheet.cell(row=4, column=2).value = total_income_bgn
        summary_sheet.cell(row=4, column=3).value = total_income_euro
        
        # Add expenses row
        summary_sheet.cell(row=5, column=1).value = "EXPENSES"
        summary_sheet.cell(row=5, column=2).value = total_expenses_bgn
        summary_sheet.cell(row=5, column=3).value = total_expenses_euro
        
        # Add total row
        summary_sheet.cell(row=6, column=1).value = "NET PROFIT/LOSS"
        summary_sheet.cell(row=6, column=1).font = Font(bold=True)
        
        net_profit_bgn = total_income_bgn - total_expenses_bgn
        net_profit_euro = total_income_euro - total_expenses_euro
        
        summary_sheet.cell(row=6, column=2).value = net_profit_bgn
        summary_sheet.cell(row=6, column=3).value = net_profit_euro
        
        # Add color coding based on net profit/loss
        profit_loss_cell_bgn = summary_sheet.cell(row=6, column=2)
        profit_loss_cell_euro = summary_sheet.cell(row=6, column=3)
        
        if net_profit_bgn > 0:
            profit_loss_cell_bgn.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        elif net_profit_bgn < 0:
            profit_loss_cell_bgn.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            
        if net_profit_euro > 0:
            profit_loss_cell_euro.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        elif net_profit_euro < 0:
            profit_loss_cell_euro.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            
        # Add borders to all summary cells
        for row in range(3, 7):
            for col in range(1, 5):
                cell = summary_sheet.cell(row=row, column=col)
                cell.border = thin_border
                if col in [2, 3]:
                    cell.alignment = Alignment(horizontal='right')
                elif col == 1:
                    cell.alignment = Alignment(horizontal='left')
    
    # Seek to the beginning of the buffer
    buffer.seek(0)
    
        # Add alerts for suspicious amount values
    alerts = []
    
    # Check for suspicious values in income and expense sheets
    for sheet_name in [income_sheet_name, expense_sheet_name]:
        sheet = writer.sheets[sheet_name]
        
        # Define columns to check based on sheet type
        amount_columns = []
        if sheet_name == income_sheet_name:
            for col_idx, col_name in enumerate(income_columns, 1):
                if any(term in col_name for term in ["NET", "VAT", "TOTAL"]):
                    amount_columns.append(col_idx)
        
        # Check all rows for suspicious values
        for row in range(2, sheet.max_row + 1):  # Start from row 2 (skip header)
            for col in amount_columns:
                cell = sheet.cell(row=row, column=col)
                
                # Skip empty cells
                if not cell.value:
                    continue
                
                try:
                    # Handle different data types
                    cell_value = cell.value
                    if isinstance(cell_value, str):
                        # Try to convert string to number
                        clean_value = cell_value.replace("$", "").replace("€", "").replace(",", ".")
                        cell_value = float(clean_value)
                    
                    # Check for suspicious values (0.0 or very small values)
                    if cell_value == 0 or (isinstance(cell_value, (int, float)) and cell_value < 0.01):
                        # Mark the cell with red background
                        cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                        
                        # Get context information for this cell
                        col_name = sheet.cell(row=1, column=col).value
                        transactor_col = income_columns.index("TRANSACTORS") + 1 if sheet_name == income_sheet_name else None
                        transactor = sheet.cell(row=row, column=transactor_col).value if transactor_col else "Unknown"
                        
                        invoice_col = income_columns.index("NUMBER OF INVOICE") + 1 if sheet_name == income_sheet_name else None
                        invoice_num = sheet.cell(row=row, column=invoice_col).value if invoice_col else "Unknown"
                        
                        # Create alert
                        alert_msg = f"REVIEW NEEDED: Suspicious amount value {cell_value} for {transactor} (Invoice: {invoice_num}, Field: {col_name})"
                        alerts.append(alert_msg)
                except (ValueError, TypeError, IndexError):
                    # Not a numeric value or column not found
                    continue
    
    return buffer.getvalue(), alerts

def export_suppliers_to_excel(suppliers_data: list) -> bytes:
    """
    Export suppliers data to Excel with only the required fields:
    Supplier Name (Transactor), Service Category, Transaction Type, and VAT Number
    
    Args:
        suppliers_data: List of supplier dictionaries
        
    Returns:
        Excel file as bytes
    """
    # Create a new list with only the required fields
    filtered_data = []
    for supplier in suppliers_data:
        # Get category - use categories list if available
        category = ''
        if 'categories' in supplier and supplier['categories']:
            category = ', '.join(supplier['categories'])
        elif 'category' in supplier and supplier['category']:
            category = supplier['category']
            
        # Get VAT number from either vat_number or tax_amount field
        vat_number = supplier.get('vat_number') or supplier.get('tax_amount', '')
            
        filtered_supplier = {
            'Supplier Name (Transactor)': supplier.get('supplier_name', ''),
            'Service Category': category,
            'Transaction Type': supplier.get('transaction_type', ''),
            'VAT Number': vat_number
        }
        filtered_data.append(filtered_supplier)
        
    df = pd.DataFrame(filtered_data)
    return export_to_excel(df, "Suppliers")

def export_invoice_data_to_excel(invoice_data: list) -> bytes:
    """
    Export invoice data to Excel with only the required fields:
    Supplier Name (Transactor), Service Category, Transaction Type, and VAT Number
    
    Args:
        invoice_data: List of invoice data dictionaries
        
    Returns:
        Excel file as bytes
    """
    # Create a new list with only the required fields
    filtered_data = []
    for invoice in invoice_data:
        # Get category - use categories list if available
        category = ''
        if 'categories' in invoice and invoice['categories']:
            category = ', '.join(invoice['categories'])
        elif 'category' in invoice and invoice['category']:
            category = invoice['category']
        elif invoice.get('raw_extraction') and invoice['raw_extraction'].get('Expense Category Account'):
            category = invoice['raw_extraction'].get('Expense Category Account')
            
        # Get VAT number from either vat_number or tax_amount field
        vat_number = invoice.get('vat_number') or invoice.get('tax_amount', '')
        
        # Get transaction type from the appropriate field
        transaction_type = invoice.get('transaction_type', '')
        if not transaction_type and invoice.get('raw_extraction'):
            transaction_type = invoice['raw_extraction'].get('Kind of Transaction', '')
            
        # Get supplier name/transactor from the appropriate field
        supplier_name = invoice.get('supplier_name', '')
        if not supplier_name and invoice.get('raw_extraction'):
            supplier_name = invoice['raw_extraction'].get('Transactor', '')
            
        filtered_invoice = {
            'Supplier Name (Transactor)': supplier_name,
            'Service Category': category,
            'Transaction Type': transaction_type,
            'VAT Number': vat_number
        }
        filtered_data.append(filtered_invoice)
        
    df = pd.DataFrame(filtered_data)
    return export_to_excel(df, "Invoice Data")

def export_monthly_financial_report(invoice_data: list) -> bytes:
    """
    Export a formatted monthly financial report focusing only on the required fields:
    Supplier Name (Transactor), Service Category, Transaction Type, and VAT Number.
    
    Args:
        invoice_data: List of invoice data dictionaries
        
    Returns:
        Excel file as bytes with formatted monthly report
    """
    # First, extract only the required fields
    filtered_data = []
    for invoice in invoice_data:
        # Get category - use categories list if available
        category = ''
        if 'categories' in invoice and invoice['categories']:
            category = ', '.join(invoice['categories'])
        elif 'category' in invoice and invoice['category']:
            category = invoice['category']
        elif invoice.get('raw_extraction') and invoice['raw_extraction'].get('Expense Category Account'):
            category = invoice['raw_extraction'].get('Expense Category Account')
            
        # Get VAT number from either vat_number or tax_amount field
        vat_number = invoice.get('vat_number') or invoice.get('tax_amount', '')
        
        # Get transaction type from the appropriate field
        transaction_type = invoice.get('transaction_type', '')
        if not transaction_type and invoice.get('raw_extraction'):
            transaction_type = invoice['raw_extraction'].get('Kind of Transaction', '')
            
        # Get supplier name/transactor from the appropriate field
        supplier_name = invoice.get('supplier_name', '')
        if not supplier_name and invoice.get('raw_extraction'):
            supplier_name = invoice['raw_extraction'].get('Transactor', '')
            
        # Keep necessary fields for the monthly report
        filtered_invoice = invoice.copy()
        filtered_invoice.update({
            'supplier_name': supplier_name,
            'category': category,
            'transaction_type': transaction_type,
            'vat_number': vat_number
        })
        filtered_data.append(filtered_invoice)
    
    # Get current month and year for the report
    current_date = datetime.now()
    month = current_date.strftime("%B")
    year = str(current_date.year)
    
    # Create the monthly report using the filtered data
    return create_monthly_financial_report(filtered_data, month, year)
    
def create_comprehensive_financial_report(invoice_data: list, bank_statement_df: Optional[pd.DataFrame] = None, month: Optional[str] = None, year: Optional[str] = None, supplier_manager=None) -> Tuple[bytes, List[str]]:
    """
    Create a comprehensive financial report combining database invoices and bank statement transactions.
    
    Args:
        invoice_data: List of invoice data dictionaries from the database
        bank_statement_df: DataFrame containing processed bank statement data (optional)
        month: Month for the report (default: current month)
        year: Year for the report (default: current year)
        supplier_manager: SupplierManager instance to get updated supplier information
        
    Returns:
        Excel file as bytes with the comprehensive report
    """
    import logging
    logger = logging.getLogger(__name__)
    
    # Set default month and year if not provided
    if month is None:
        month = datetime.now().strftime("%B")
    if year is None:
        year = str(datetime.now().year)
    
    logger.info(f"Creating comprehensive financial report for {month} {year}")
    
    # Convert month name to month number if needed
    month_name_to_num = {
        "January": 1, "February": 2, "March": 3, "April": 4, "May": 5, "June": 6,
        "July": 7, "August": 8, "September": 9, "October": 10, "November": 11, "December": 12
    }
    month_num = month_name_to_num.get(month, datetime.now().month)
    
    # Create a buffer to store the Excel file
    buffer = io.BytesIO()
    
    # Prepare data for reporting
    expense_data = []
    income_data = []
    
    # Import SupplierManager if not provided
    if supplier_manager is None:
        try:
            from supplier_manager import SupplierManager
            supplier_manager = SupplierManager()
            logger.info(f"Created new SupplierManager instance for report")
        except ImportError:
            logger.warning("Could not import SupplierManager - will use invoice data as-is")
    
    # Track the suppliers we've already processed to avoid duplicates
    processed_supplier_invoices = set()
    
    # Process invoice data from database - include ALL database records for this month/year
    logger.info(f"Processing {len(invoice_data)} invoices for report")
    
    # Check if we have any Upwork invoices for debugging
    upwork_invoices = [inv for inv in invoice_data if inv.get('transactor') and 'upwork' in inv.get('transactor').lower()]
    logger.info(f"Found {len(upwork_invoices)} Upwork invoices to process")
    
    for invoice in invoice_data:
        # Parse the date to check if it's in the requested month/year
        invoice_date = invoice.get('date')
        
        # Skip if no date and we're not sure it belongs to this month
        if not invoice_date:
            logger.debug(f"Invoice missing date, using default: {invoice}")
            # If no date, we'll include it anyway since we want all database records
            invoice_date = f"{year}-{month_num:02d}-01"  # Set to 1st of selected month as default
            
        # Handle different date formats
        invoice_year = None
        invoice_month = None
        
        if isinstance(invoice_date, str):
            try:
                # Handle YYYY-MM-DD format (SQL standard)
                if '-' in invoice_date:
                    date_parts = invoice_date.split('-')
                    if len(date_parts) >= 2:
                        invoice_year = int(date_parts[0])
                        invoice_month = int(date_parts[1])
                # Handle DD/MM/YYYY format (European)
                elif '/' in invoice_date:
                    date_parts = invoice_date.split('/')
                    if len(date_parts) >= 3:
                        invoice_day = int(date_parts[0])
                        invoice_month = int(date_parts[1])
                        invoice_year = int(date_parts[2])
                    elif len(date_parts) == 2:
                        invoice_day = int(date_parts[0])
                        invoice_month = int(date_parts[1])
                        invoice_year = int(year)  # Use the selected year
                # Handle MM/DD/YYYY format (US)
                elif invoice_date.count('/') == 2 and len(invoice_date.split('/')[0]) <= 2:
                    date_parts = invoice_date.split('/')
                    invoice_month = int(date_parts[0])
                    invoice_day = int(date_parts[1])
                    invoice_year = int(date_parts[2])
                # Handle text month format like "January 2023"
                elif len(invoice_date.split()) == 2:
                    month_str, year_str = invoice_date.split()
                    month_map = {
                        "january": 1, "february": 2, "march": 3, "april": 4,
                        "may": 5, "june": 6, "july": 7, "august": 8,
                        "september": 9, "october": 10, "november": 11, "december": 12
                    }
                    invoice_month = month_map.get(month_str.lower(), None)
                    invoice_year = int(year_str)
                else:
                    # Just use what's provided in the function arguments
                    invoice_year = int(year)
                    invoice_month = int(month_num)
            except (ValueError, IndexError):
                # On error, use the selected month/year
                invoice_year = int(year)
                invoice_month = int(month_num)
        
        # If we couldn't determine the date, use the selected month/year
        if invoice_year is None or invoice_month is None:
            invoice_year = int(year)
            invoice_month = int(month_num)
            
        # Skip if not in the requested month/year
        if str(invoice_year) != str(year) or invoice_month != int(month_num):
            continue
            
        # Check if this is an income or expense transaction
        is_income = invoice.get('is_income', False)
        
        # Get transactor/supplier name first (needed to look up supplier info)
        transactor = invoice.get('transactor', '')
        if not transactor:
            transactor = invoice.get('supplier_name', '')
            
        # Log for debugging Upwork transactions
        if 'upwork' in transactor.lower():
            logger.info(f"⚠️ Processing Upwork invoice: {invoice.get('invoice_number')} - {invoice.get('description')}")
        
        # If we have a supplier_manager, look up the supplier to get latest info
        supplier_data = None
        if supplier_manager and transactor:
            try:
                supplier_data = supplier_manager.get_supplier(transactor)
                logger.info(f"Found supplier data for '{transactor}' in database")
            except Exception as e:
                logger.warning(f"Error retrieving supplier data: {e}")
        
        # Get transaction type - prioritize data from supplier database
        transaction_type = ''
        if supplier_data and supplier_data.get('transaction_type'):
            # Use supplier database transaction type (most reliable)
            transaction_type = supplier_data.get('transaction_type')
            logger.info(f"Using supplier database transaction type: {transaction_type}")
        elif invoice.get('transaction_type'):
            # Use invoice transaction type as fallback
            transaction_type = invoice.get('transaction_type')
        else:
            # Default if nothing else available
            transaction_type = 'INCOME' if is_income else 'EXPENSES with VAT'
            
        # Get expense category - prioritize data from supplier database
        category = ''
        if supplier_data:
            # Priority 1: Use categories from supplier database
            supplier_categories = supplier_data.get('categories', [])
            if supplier_categories and isinstance(supplier_categories, list) and len(supplier_categories) > 0:
                # Make sure we have a valid list with non-empty values before joining
                categories_to_join = [str(c) for c in supplier_categories if c]
                if categories_to_join:
                    category = ', '.join(categories_to_join)
                    logger.info(f"Using supplier database categories: {category}")
            # Priority 2: Use single category from supplier database
            elif supplier_data.get('category'):
                category = supplier_data.get('category')
                logger.info(f"Using supplier database category: {category}")
        
        # Only if not found in supplier database, try invoice data
        if not category:
            invoice_categories = invoice.get('categories', [])
            if invoice_categories and isinstance(invoice_categories, list) and len(invoice_categories) > 0:
                # Make sure we have a valid list with non-empty values before joining
                categories_to_join = [str(c) for c in invoice_categories if c]
                if categories_to_join:
                    category = ', '.join(categories_to_join)
            elif invoice.get('category'):
                category = invoice.get('category')
            elif invoice.get('expense_category'):
                category = invoice.get('expense_category')
                
        # Ensure we always have a string
        if not category:
            category = ""
        
        # Get account description (Kind of Transaction) - might be in different fields
        account_description = invoice.get('account_description', '')
        if not account_description:
            account_description = transaction_type
            
        # Handle amounts
        amount = invoice.get('amount', 0)
        
        # Handle VAT - if it's 0, None, empty string, or just whitespace, set to empty string
        vat = invoice.get('vat', '')
        if vat == 0 or vat is None or (isinstance(vat, str) and vat.strip() == ''):
            vat = ""
            
        total_bgn = invoice.get('total_bgn', 0)
        total_euro = invoice.get('total_euro', 0)
        
        # Format the date to match the exact format from the example DD/MM/YYYY
        formatted_date = invoice_date
        if isinstance(invoice_date, str) and '-' in invoice_date:
            # Convert YYYY-MM-DD to DD/MM/YYYY
            try:
                parts = invoice_date.split('-')
                if len(parts) == 3:
                    formatted_date = f"{parts[2]}/{parts[1]}/{parts[0]}"
            except:
                pass
                
        # Determine payment method
        payment_method = 'supplier'
        
        # Special handling for Upwork - must be included in comprehensive report
        if transactor and 'upwork' in transactor.lower():
            # Check if description contains "paypal" to determine payment method
            description = invoice.get('description', '')
            notes = invoice.get('notes', '')
            
            if description and 'paypal' in description.lower():
                payment_method = 'PAYPAL'
                logger.info(f"Setting payment method to PAYPAL because 'paypal' was found in description for Upwork invoice")
            elif notes and 'paypal' in notes.lower():
                payment_method = 'PAYPAL'
                logger.info(f"Setting payment method to PAYPAL because 'paypal' was found in notes for Upwork invoice")
            
            logger.info(f"Processing Upwork transaction with payment method: {payment_method}")
        else:
            # For other invoice data from database, use standard payment method
            logger.info(f"Using '{payment_method}' as payment method for invoice from database")
        
        # Handle currency according to the exact format in the example
        currency = invoice.get('currency', 'EUR')
        
        # Format according to the new requirements:
        # - For USD ($): Show in Amount column with $ prefix, convert to EUR in Total (EUR), empty BGN, currency="EU"
        # - For BGN: Show only in Total (BGN), empty Amount, empty EUR, currency="BG"
        # - For EUR: Show only in Total (EUR), empty Amount, empty BGN, currency="EU"
        if currency in ['USD', '$']:
            # For USD, show amount in Amount column with $ prefix, convert to EUR
            display_currency = 'EU'  # USD amounts show "EU" in currency column
            
            # Format amount with $ prefix for USD
            raw_amount = invoice.get('amount', 0)
            if isinstance(raw_amount, (int, float)):
                # Format with $ prefix
                amount = f"${raw_amount}"
            else:
                # If it's already a string, make sure it has $ prefix
                amount = raw_amount if str(raw_amount).startswith('$') else f"${raw_amount}"
            
            # Set the values according to requirements
            # VAT handling is already done above
            total_bgn = ""  # Empty for USD
            
            # Add currency conversion alert icon with warning triangle for USD to EUR conversions
            # This alerts the user that the EUR value might differ due to exchange rate fluctuations
            raw_total_euro = invoice.get('total_euro', 0)
            if isinstance(raw_total_euro, (int, float)) and raw_total_euro > 0:
                # For Excel exports, we need to use a special prefix that works in Excel
                # We'll use a colored warning symbol (⚠) as the indicator
                total_euro = f"⚠ {raw_total_euro}"
            else:
                total_euro = raw_total_euro  # Keep as is if already formatted or zero
            
        elif currency in ['BGN', 'BG', 'G']:
            # For BGN, show only in Total (BGN), empty Amount and Total (EUR)
            display_currency = 'BG'  # BGN amounts show "BG" in currency column
            
            # Set the values according to requirements
            amount = ""  # Empty Amount column for BGN
            # VAT handling is already done above
            total_bgn = invoice.get('total_bgn', 0)  # BGN total only
            total_euro = ""  # Empty for BGN
            
        else:
            # For EUR, show only in Total (EUR), empty Amount and Total (BGN)
            display_currency = 'EU'  # EUR amounts show "EU" in currency column
            
            # Set the values according to requirements
            amount = ""  # Empty Amount column for EUR
            # VAT handling is already done above
            total_bgn = ""  # Empty for EUR
            total_euro = invoice.get('total_euro', 0)  # Show in Total (EUR) column
            
        # Get description from invoice, fallback to notes if available
        # Make sure not to display "0" for empty descriptions - keep them truly empty
        description = invoice.get('description', '')
        if not description and invoice.get('notes'):
            description = invoice.get('notes')
        # Make sure empty values are actually empty strings, not "0"
        if description == 0 or description == "0":
            description = ""
            
        if is_income:
            income_item = {
                "Date": formatted_date,
                "Invoice #": invoice.get('invoice_number', ''),
                "Transactor": transactor,
                "Description": description,
                "Kind of Transaction": transaction_type,
                "Amount": amount,
                "VAT": vat,
                "Total (BGN)": total_bgn,
                "Total (EUR)": total_euro,
                "Currency": display_currency,
                "Payment Method": payment_method
            }
            income_data.append(income_item)
        else:
            # For expenses, format exactly like the example format you provided
            expense_item = {
                "Date": formatted_date,
                "Invoice #": invoice.get('invoice_number', ''),
                "Transactor": transactor,
                "Expense Category": category,
                "Kind of Transaction": transaction_type,
                "Description": description,  # Use the processed description
                "Amount": amount,
                "VAT": vat,
                "Total (BGN)": total_bgn,
                "Total (EUR)": total_euro,
                "Currency": display_currency,
                "Payment Method": payment_method
            }
            expense_data.append(expense_item)
    
    # Process bank statement data if provided
    bank_data = []
    if bank_statement_df is not None:
        logger.info(f"Processing bank statement DataFrame with {len(bank_statement_df)} rows")
        logger.info(f"Bank statement columns: {list(bank_statement_df.columns)}")
        logger.info(f"Bank statement dtypes: {bank_statement_df.dtypes}")
        
        # Print first two rows to debug what data format we're dealing with
        if len(bank_statement_df) > 0:
            try:
                first_row = bank_statement_df.iloc[0]
                logger.info(f"First row values: {first_row.to_dict()}")
                if len(bank_statement_df) > 1:
                    second_row = bank_statement_df.iloc[1]
                    logger.info(f"Second row values: {second_row.to_dict()}")
            except Exception as e:
                logger.error(f"Error examining rows: {str(e)}")
        
        # Process both processed and raw bank statements
        for index, row in bank_statement_df.iterrows():
            logger.info(f"Processing bank statement row {index}")
            row_data = row.to_dict()
            logger.info(f"Row data sample: {str(row_data)[:200]}...")
            
            # 1. First, try to find a date column and value
            date_col = None
            date_value = None
            
            # Try to find a date column by name - specifically check for both "Date" (standard) 
            # and "S/N", "Reference Number" columns (bank format with header on row 6)
            for col in row.index:
                col_str = str(col).lower()
                
                # Check for standard date column
                if col_str == 'date':
                    date_col = col
                    date_value = row.get(col)
                    logger.info(f"Found exact date column: {date_col} with value: {date_value}")
                    break
                
                # Check for partial date column name
                if 'date' in col_str and 'value' not in col_str:
                    date_col = col
                    date_value = row.get(col)
                    logger.info(f"Found date column: {date_col} with value: {date_value}")
                    break
            
            # If no explicit date column, look for date-like values
            if date_col is None:
                logger.info("No explicit date column found, checking column content")
                for col in row.index:
                    value = row.get(col)
                    
                    # First check for datetime objects (pandas automatically converts Excel dates to datetime)
                    if isinstance(value, datetime):
                        date_col = col
                        date_value = value
                        logger.info(f"Found datetime object in column '{col}': {value}")
                        break
                        
                    # Check if the value looks like a date string
                    elif isinstance(value, str) and ('/' in value or '-' in value):
                        # Simple check to see if it has digits and separators like a date
                        if any(c.isdigit() for c in value):
                            date_col = col
                            date_value = value
                            logger.info(f"Found potential date value '{value}' in column '{col}'")
                            break
            
            # Skip if no date found
            if date_col is None:
                logger.info("Skipping row without date")
                continue
            
            # 2. Parse the date value
            if date_value is not None:
                # Check if it's already a datetime object
                if isinstance(date_value, datetime):
                    bank_day = date_value.day
                    bank_month = date_value.month
                    bank_year = date_value.year
                    logger.info(f"Using datetime object directly: day={bank_day}, month={bank_month}, year={bank_year}")
                else:
                    date_str = str(date_value)
                    logger.info(f"Parsing date from: {date_str}")
                    
                    try:
                        if '/' in date_str:
                            date_parts = date_str.split('/')
                            bank_day = int(date_parts[0]) if len(date_parts) > 0 else 1
                            bank_month = int(date_parts[1]) if len(date_parts) > 1 else 1
                            bank_year = int(date_parts[2]) if len(date_parts) > 2 else int(year)
                            logger.info(f"Parsed date parts from '/': day={bank_day}, month={bank_month}, year={bank_year}")
                        elif '-' in date_str:
                            date_parts = date_str.split('-')
                            bank_year = int(date_parts[0]) if len(date_parts) > 0 else int(year)
                            bank_month = int(date_parts[1]) if len(date_parts) > 1 else 1
                            bank_day = int(date_parts[2]) if len(date_parts) > 2 else 1
                            logger.info(f"Parsed date parts from '-': day={bank_day}, month={bank_month}, year={bank_year}")
                        # Handle Excel date format (could be a float or int)
                        elif isinstance(date_value, (int, float)):
                            # Excel dates are days since 1899-12-30 (for Windows)
                            excel_epoch = datetime(1899, 12, 30)
                            try:
                                date_obj = excel_epoch + timedelta(days=int(date_value))
                                bank_day = date_obj.day
                                bank_month = date_obj.month
                                bank_year = date_obj.year
                                logger.info(f"Parsed date from Excel number: day={bank_day}, month={bank_month}, year={bank_year}")
                            except Exception as e:
                                logger.error(f"Error parsing Excel date: {str(e)}")
                                continue
                        else:
                            logger.info(f"Unrecognized date format: {date_str}")
                            continue
                    except (ValueError, IndexError) as e:
                        logger.error(f"Error parsing date {date_str}: {str(e)}")
                        continue
                
                # Skip if not in the requested month/year
                if bank_year != int(year) or bank_month != int(month_num):
                    logger.info(f"Skipping transaction not in requested month/year: {bank_month}/{bank_year}")
                    continue
            else:
                logger.info("No date value found, skipping row")
                continue
            
            # 3. Find the transaction amount
            amount = None
            
            # First look for an explicit Amount column
            if 'Amount' in row:
                amount = row.get('Amount')
                logger.info(f"Found amount in 'Amount' column: {amount}")
            
            # If still no amount, try to find a column that might contain amount data
            if amount is None:
                for col in row.index:
                    col_str = str(col).lower()
                    # Look for amount-related column names
                    if 'amount' in col_str or 'sum' in col_str or 'total' in col_str or 'value' in col_str:
                        amount = row.get(col)
                        logger.info(f"Found amount in column '{col}': {amount}")
                        break
            
            # If we found something, ensure it's a numeric value
            if amount is not None:
                try:
                    # Handle amounts with European number format (comma as decimal)
                    if isinstance(amount, str):
                        # Check if it has commas or dots
                        if ',' in amount and '.' in amount:
                            # European format with thousands separators
                            amount = amount.replace('.', '')  # Remove thousand separator
                            amount = amount.replace(',', '.')  # Convert decimal separator
                        elif ',' in amount:
                            # Might be European format with comma as decimal
                            amount = amount.replace(',', '.')
                    
                    # Convert to float
                    amount = float(amount)
                    logger.info(f"Converted amount to numeric value: {amount}")
                except (ValueError, TypeError) as e:
                    logger.error(f"Error converting amount {amount} to number: {str(e)}")
                    amount = None
            
            # Skip if no valid amount
            if amount is None:
                logger.info("No valid amount found, skipping")
                continue
                
            # 4. Determine transaction sign
            amount_sign = ''
            
            # Check for explicit Amount Sign column from processed statements
            if 'Amount Sign' in row:
                amount_sign = str(row.get('Amount Sign', '')).strip().upper()
                logger.info(f"Found sign in 'Amount Sign' column: '{amount_sign}'")
            
            # Try to find a sign column in raw statements
            if not amount_sign:
                for col in row.index:
                    col_str = str(col).lower()
                    if 'sign' in col_str:
                        sign_value = str(row.get(col, '')).strip().upper()
                        if sign_value:
                            amount_sign = sign_value
                            logger.info(f"Found sign in '{col}' column: '{amount_sign}'")
                            break
            
            # Check fields for "D" or "C" indicators
            if not amount_sign:
                for col in row.index:
                    value = str(row.get(col, '')).strip().upper()
                    # If the value is just "D" or "C" (or close to it)
                    if value in ['D', 'C', 'DR', 'CR', 'DEBIT', 'CREDIT']:
                        amount_sign = 'D' if value in ['D', 'DR', 'DEBIT'] else 'C'
                        logger.info(f"Found sign indicator in column '{col}': '{value}' -> '{amount_sign}'")
                        break
            
            # If still no sign, try to infer from amount
            if not amount_sign:
                # For bank statements, negative amounts are typically debits
                amount_sign = 'D' if amount < 0 else 'C'
                logger.info(f"Inferred sign from amount {amount}: '{amount_sign}'")
            
            # STRICT CHECK: Skip if not a debit (D) transaction as per user's requirement
            if amount_sign != 'D':
                logger.info(f"Skipping non-debit transaction with sign: '{amount_sign}'")
                continue
                
            logger.info(f"Processing debit transaction with sign: '{amount_sign}'")
            
            # Make negative amounts positive for better display in reports
            if amount < 0:
                amount = abs(amount)
                logger.info(f"Converting negative amount to positive: {amount}")
            
            # One more integrity check - amount should not be very small (near zero)
            if abs(amount) < 0.01:
                logger.warning(f"Suspiciously small amount detected: {amount} - skipping")
                continue
            
            # For bank statements with "D" sign entries:
            # 1. Format specifically for comprehensive report:
            #    - For bank fees: Kind of Transaction="BANK FEES", Transactor="BANK"
            #    - For other payments: Kind of Transaction="PAYMENT", Transactor=payee name
            #    - Payment Mode should be BANK or PAYPAL depending on transaction
            
            # Determine transaction type based on description and amount
            # Default to PAYMENT per user requirements
            transaction_type = 'PAYMENT'
            # Extract transactor, reference and description from appropriate columns
            transactor = row.get('Payee', '')
            reference = None
            description = row.get('Description', '')
            
            # Look for Reference Number column (exact or partial match)
            for col in row.index:
                col_str = str(col).lower()
                if col_str == 'reference number' or 'reference' in col_str:
                    reference = row.get(col)
                    logger.info(f"Found reference in {col} column: {reference}")
                    break
                    
            # If no reference found, use any unnamed column that might contain it
            if reference is None:
                for col in row.index:
                    if 'unnamed' in str(col).lower() and row.get(col) and isinstance(row.get(col), str):
                        reference = row.get(col)
                        logger.info(f"Found possible reference in {col} column: {reference}")
                        break
                        
            # If still no reference, handle as empty string
            if reference is None:
                reference = ''
            
            # Look for bank fees in reference, payee or description
            # (common pattern has "BANK FEES", "BANK FEE", or transaction ≤17)
            is_bank_fee = False
            
            # Check reference for bank fee indicators
            if reference and ("BANK FEE" in str(reference).upper() or "BANK FEES" in str(reference).upper()):
                is_bank_fee = True
            
            # Check payee for bank fee indicators
            elif transactor and (
                "FEE" in str(transactor).upper() and "BANK" in str(transactor).upper() or
                "BANK FEE" in str(transactor).upper() or 
                "BANK FEES" in str(transactor).upper()
            ):
                is_bank_fee = True
                
            # Check description for bank fee indicators
            elif description and (
                "BANK FEE" in str(description).upper() or 
                "BANK FEES" in str(description).upper() or
                "COMMISSION" in str(description).upper()
            ):
                is_bank_fee = True
                
            # Check for small amount (typical bank fees)
            elif abs(float(amount)) <= 17:  # Small amounts are typically bank fees
                is_bank_fee = True
            
            # Set transaction_type and transactor for bank fees  
            if is_bank_fee:
                transaction_type = 'BANK FEES'
                transactor = 'BANK'  # Default transactor for bank fees
            
            # If no payee was found for a payment, use reference or placeholder
            if transaction_type == 'PAYMENT' and not transactor:
                if reference:
                    transactor = reference
                else:
                    transactor = 'Unknown Payee'
            
            # Format the date as DD/MM/YYYY
            formatted_date = f"{bank_day}/{bank_month}/{bank_year}"
            
            # Determine payment method (BANK or PAYPAL)
            payment_method = 'BANK'
            
            # Check for PayPal in reference, transactor, or description
            if ("PAYPAL" in str(reference).upper() or 
                "PAYPAL" in str(transactor).upper() or 
                (description and "PAYPAL" in str(description).upper())):
                payment_method = 'PAYPAL'
                logger.info(f"Setting payment method to PAYPAL because 'paypal' was found in transaction data")
                
            # Special handling for Upwork transactions
            if (transactor and "UPWORK" in str(transactor).upper() or 
                (description and "UPWORK" in str(description).upper()) or
                (reference and "UPWORK" in str(reference).upper())):
                # For Upwork, check if description contains "paypal" to determine payment method
                if (description and "PAYPAL" in str(description).upper()):
                    payment_method = 'PAYPAL'
                    logger.info(f"Setting payment method to PAYPAL for Upwork transaction with PayPal in description")
            
            # Format bank transaction exactly as requested with:
            # Bank fees format: [Date, BANK FEES, BANK, , , , , , , Amount, EU, BANK]
            # Payment format: [Date, PAYMENT, Supplier Name, , , , , , , Amount, EU, BANK/PAYPAL]
            
            # For bank transactions, we follow the EUR pattern: Empty Amount and BGN, show in EUR column
            bank_item = {
                "Date": formatted_date,
                "Kind of Transaction": transaction_type,  # BANK FEES or PAYMENT
                "Transactor": transactor,  # BANK for bank fees, payee name for other payments
                "Expense Category": "",  # Empty as requested
                "Description": "",  # Empty as requested
                "Invoice #": "",  # Empty as requested (Number of Invoice) 
                "Amount": "",  # Empty as per EUR format
                "VAT": "",  # Empty as requested
                "Total (BGN)": "",  # Empty as per EUR format
                "Total (EUR)": float(amount),  # Add amount to Euro column for all bank transactions
                "Currency": "EU",  # Always EU for bank transactions (per user requirement)
                "Payment Method": payment_method  # BANK or PAYPAL
            }
            bank_data.append(bank_item)
    
    # Combine expense data from database with bank transaction data
    # This ensures we have a comprehensive list of all expenses
    combined_expense_data = expense_data.copy()
    
    # Add bank data to combined expense data - but only if it's not already in the database
    if bank_data:
        for bank_item in bank_data:
            # Check if this bank transaction matches any known transaction in the database
            is_duplicate = False
            for expense_item in expense_data:
                # Compare by date, transactor, and amount (could be in different columns based on currency)
                # For bank items, the amount is in Total (EUR) and Amount is empty
                bank_date = bank_item.get("Date", "")
                bank_transactor = bank_item.get("Transactor", "")
                bank_amount_eur = bank_item.get("Total (EUR)", 0)
                
                expense_date = expense_item.get("Date", "")
                expense_transactor = expense_item.get("Transactor", "")
                
                # Get amount from appropriate column based on currency
                expense_currency = expense_item.get("Currency", "")
                if expense_currency == "EU":
                    # For both USD and EUR, the amount will be in Total (EUR)
                    expense_amount = expense_item.get("Total (EUR)", 0)
                elif expense_currency == "BG":
                    # For BGN, the amount will be in Total (BGN)
                    expense_amount = expense_item.get("Total (BGN)", 0)
                else:  # Fallback for any other currencies
                    # Try to find a non-empty amount value in any column
                    if expense_item.get("Amount"):
                        expense_amount = expense_item.get("Amount", 0)
                    elif expense_item.get("Total (EUR)"):
                        expense_amount = expense_item.get("Total (EUR)", 0)
                    elif expense_item.get("Total (BGN)"):
                        expense_amount = expense_item.get("Total (BGN)", 0)
                    else:
                        expense_amount = 0
                
                # Check if this looks like the same transaction
                if (
                    bank_date == expense_date and
                    bank_transactor == expense_transactor and
                    (
                        # Either the EUR amounts match
                        (bank_amount_eur and expense_amount and 
                         abs(float(bank_amount_eur) - float(expense_amount)) < 0.01)
                    )
                ):
                    is_duplicate = True
                    break
            
            # If it's not a duplicate, add it to the combined data
            if not is_duplicate:
                combined_expense_data.append(bank_item)
    
    # Create Excel workbook
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        workbook = writer.book
        
        # Create combined expense sheet for the comprehensive report
        expense_sheet_name = "EXPENSES"
        
        # Create the columns in the exact order from your provided example
        # Description column is kept prominently near the beginning for better visibility
        column_order = [
            "Date", 
            "Invoice #", 
            "Transactor", 
            "Description",  # Description moved to more prominent position
            "Expense Category", 
            "Kind of Transaction", 
            "Amount", 
            "VAT",
            "Total (BGN)",
            "Total (EUR)",
            "Currency", 
            "Payment Method"
        ]
        
        # Create DataFrame with the specified column order
        expense_df = pd.DataFrame(combined_expense_data)
        
        # Reorganize columns to match the example format
        ordered_cols = [col for col in column_order if col in expense_df.columns]
        missing_cols = [col for col in column_order if col not in expense_df.columns]
        
        # Add any missing columns with empty values
        for col in missing_cols:
            expense_df[col] = ""
            
        # Reorder columns to match the expected format
        expense_df = expense_df[ordered_cols + [col for col in expense_df.columns if col not in ordered_cols]]
        
        if not expense_df.empty:
            # Sort by date
            if "Date" in expense_df.columns:
                try:
                    # Try to parse dates for sorting (may fail due to format differences)
                    expense_df = expense_df.sort_values(by="Date")
                except:
                    pass
            
            expense_df.to_excel(writer, sheet_name=expense_sheet_name, index=False, columns=ordered_cols)
        else:
            pd.DataFrame(columns=column_order).to_excel(writer, sheet_name=expense_sheet_name, index=False)
            
        # Format expense sheet
        expense_sheet = writer.sheets[expense_sheet_name]
        header_fill = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
        
        # Format headers
        for col_num, column in enumerate(ordered_cols, 1):
            cell = expense_sheet.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            
            # Set column width
            column_letter = get_column_letter(col_num)
            if column in ['Date', 'Currency']:
                expense_sheet.column_dimensions[column_letter].width = 15
            elif column in ['Kind of Transaction', 'Transactor', 'Expense Category', 'Description']:
                expense_sheet.column_dimensions[column_letter].width = 25
            elif column in ['Invoice #', 'Payment Method']:
                expense_sheet.column_dimensions[column_letter].width = 20
            else:  # Numeric columns
                expense_sheet.column_dimensions[column_letter].width = 12
        
        # Create income sheet
        income_sheet_name = "INCOME"
        income_df = pd.DataFrame(income_data)
        
        # Define income column order based on the format of the example
        income_column_order = [
            "Date", 
            "Invoice #", 
            "Transactor", 
            "Description", 
            "Kind of Transaction", 
            "Amount", 
            "VAT",
            "Total (BGN)",
            "Total (EUR)",
            "Currency", 
            "Payment Method"
        ]
        
        # Reorganize income columns to match expected format
        income_ordered_cols = [col for col in income_column_order if col in income_df.columns]
        income_missing_cols = [col for col in income_column_order if col not in income_df.columns]
        
        # Add any missing columns with empty values
        for col in income_missing_cols:
            income_df[col] = ""
            
        # Reorder columns to match the expected format
        income_df = income_df[income_ordered_cols + [col for col in income_df.columns if col not in income_ordered_cols]]
        
        if not income_df.empty:
            # Sort by date
            if "Date" in income_df.columns:
                try:
                    income_df = income_df.sort_values(by="Date")
                except:
                    pass
            income_df.to_excel(writer, sheet_name=income_sheet_name, index=False, columns=income_column_order)
        else:
            pd.DataFrame(columns=income_column_order).to_excel(writer, sheet_name=income_sheet_name, index=False)
        
        # Format income sheet
        income_sheet = writer.sheets[income_sheet_name]
        income_header_fill = PatternFill(start_color="D1F2EB", end_color="D1F2EB", fill_type="solid")
        
        # Format headers
        for col_num, column in enumerate(income_column_order, 1):
            if col_num <= len(income_column_order):  # Make sure column exists
                cell = income_sheet.cell(row=1, column=col_num)
                cell.font = Font(bold=True)
                cell.fill = income_header_fill
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
                
                # Set column width
                column_letter = get_column_letter(col_num)
                if column in ['Date', 'Currency']:
                    income_sheet.column_dimensions[column_letter].width = 15
                elif column in ['Kind of Transaction', 'Transactor', 'Description']:
                    income_sheet.column_dimensions[column_letter].width = 25
                elif column in ['Invoice #', 'Payment Method']:
                    income_sheet.column_dimensions[column_letter].width = 20
                else:  # Numeric columns
                    income_sheet.column_dimensions[column_letter].width = 12
        
        # Create summary sheet
        summary_sheet_name = "SUMMARY"
        summary_sheet = workbook.create_sheet(summary_sheet_name)
        
        # Format SUMMARY sheet
        summary_sheet.cell(row=1, column=1).value = f"FINANCIAL SUMMARY: {month.upper()} {year}"
        summary_sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
        summary_sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')
        summary_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        
        # Add table headers
        summary_sheet.cell(row=3, column=1).value = "Category"
        summary_sheet.cell(row=3, column=2).value = "BGN"
        summary_sheet.cell(row=3, column=3).value = "Euro"
        summary_sheet.cell(row=3, column=4).value = "Notes"
        
        # Format the headers
        header_fill = PatternFill(start_color="E6F2F5", end_color="E6F2F5", fill_type="solid")
        for col in range(1, 5):
            cell = summary_sheet.cell(row=3, column=col)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Set column widths
        summary_sheet.column_dimensions['A'].width = 25
        summary_sheet.column_dimensions['B'].width = 15
        summary_sheet.column_dimensions['C'].width = 15
        summary_sheet.column_dimensions['D'].width = 25
        
        # Calculate totals
        total_expense_bgn = sum(item.get('Total BGN', 0) for item in expense_data)
        total_expense_euro = sum(item.get('Total EUR', 0) for item in expense_data)
        total_income_bgn = sum(item.get('Total BGN', 0) for item in income_data)
        total_income_euro = sum(item.get('Total EUR', 0) for item in income_data)
        
        # Add income summary row
        summary_sheet.cell(row=4, column=1).value = "TOTAL INCOME"
        summary_sheet.cell(row=4, column=2).value = total_income_bgn
        summary_sheet.cell(row=4, column=3).value = total_income_euro
        
        # Add expense summary row
        summary_sheet.cell(row=5, column=1).value = "TOTAL EXPENSES"
        summary_sheet.cell(row=5, column=2).value = total_expense_bgn
        summary_sheet.cell(row=5, column=3).value = total_expense_euro
        
        # Add net profit/loss row
        summary_sheet.cell(row=7, column=1).value = "NET PROFIT/LOSS"
        summary_sheet.cell(row=7, column=1).font = Font(bold=True)
        net_profit_bgn = total_income_bgn - total_expense_bgn
        net_profit_euro = total_income_euro - total_expense_euro
        summary_sheet.cell(row=7, column=2).value = net_profit_bgn
        summary_sheet.cell(row=7, column=3).value = net_profit_euro
        
        # Add color coding based on profit/loss
        profit_loss_fill_positive = PatternFill(start_color="D1F2EB", end_color="D1F2EB", fill_type="solid")
        profit_loss_fill_negative = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
        
        # Format profit/loss cells
        if net_profit_bgn >= 0:
            summary_sheet.cell(row=7, column=2).fill = profit_loss_fill_positive
        else:
            summary_sheet.cell(row=7, column=2).fill = profit_loss_fill_negative
            
        if net_profit_euro >= 0:
            summary_sheet.cell(row=7, column=3).fill = profit_loss_fill_positive
        else:
            summary_sheet.cell(row=7, column=3).fill = profit_loss_fill_negative
    
    # Set the workbook's active sheet to the first sheet
    workbook.active = 0
    
    # Add an alerts sheet for suspicious values
    alerts_sheet_name = "DATA ALERTS"
    alerts_sheet = workbook.create_sheet(alerts_sheet_name)
    
    # Track alerts for the report
    alerts = []
    
    # Check for suspicious amount values
    for sheet_name in [expense_sheet_name, income_sheet_name]:
        sheet = writer.sheets[sheet_name]
        # Determine which columns have amount data
        amount_columns = []
        if sheet_name == expense_sheet_name:
            # Look for amount columns in the expense sheet
            for col_num, col_name in enumerate(expense_df.columns, 1):
                if any(term in col_name for term in ['Amount', 'Total']):
                    amount_columns.append(col_num)
        else:
            # Look for amount columns in the income sheet
            for col_num, col_name in enumerate(income_df.columns, 1):
                if any(term in col_name for term in ['AMOUNT', 'NET', 'TOTAL']):
                    amount_columns.append(col_num)
        
        # Check all rows in the sheet for suspicious values
        for row in range(2, sheet.max_row + 1):  # Start from row 2 (skip header)
            for col in amount_columns:
                cell = sheet.cell(row=row, column=col)
                cell_value = cell.value
                
                # Convert cell value to float if possible
                try:
                    # Handle various formats
                    if isinstance(cell_value, str):
                        # Remove currency symbols and handle comma decimal separators
                        clean_value = cell_value.replace('$', '').replace('€', '').replace('лв', '')
                        clean_value = clean_value.strip()
                        if ',' in clean_value and '.' not in clean_value:
                            clean_value = clean_value.replace(',', '.')
                        
                        # Try to convert to float
                        if clean_value:
                            cell_value = float(clean_value)
                        else:
                            continue  # Skip empty values
                    
                    # Check for suspicious values (0.0 or very small values)
                    if cell_value == 0 or (isinstance(cell_value, (int, float)) and cell_value < 0.01):
                        # Mark cell with red background
                        cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                        
                        # Get row context for the alert
                        row_data = {}
                        for c in range(1, sheet.max_column + 1):
                            header = sheet.cell(row=1, column=c).value
                            value = sheet.cell(row=row, column=c).value
                            row_data[header] = value
                        
                        # Create alert message
                        col_name = sheet.cell(row=1, column=col).value
                        transactor = row_data.get('Transactor', '') or row_data.get('TRANSACTORS', '')
                        invoice_num = row_data.get('Invoice #', '') or row_data.get('NUMBER OF INVOICE', '')
                        
                        alert_msg = f"REVIEW NEEDED: Suspicious amount value '{cell_value}' for {transactor} (Invoice: {invoice_num}, Column: {col_name})"
                        alerts.append(alert_msg)
                        
                        # Add to alerts sheet
                        next_row = alerts_sheet.max_row + 1
                        alerts_sheet.cell(row=next_row, column=1).value = sheet_name
                        alerts_sheet.cell(row=next_row, column=2).value = transactor
                        alerts_sheet.cell(row=next_row, column=3).value = invoice_num
                        alerts_sheet.cell(row=next_row, column=4).value = col_name
                        alerts_sheet.cell(row=next_row, column=5).value = cell_value
                        alerts_sheet.cell(row=next_row, column=6).value = "Review needed - suspicious value"
                except (ValueError, TypeError):
                    # Not a numeric value, skip
                    continue
    
    # Format the alerts sheet if we have any alerts
    if alerts:
        # Add headers
        alerts_sheet.cell(row=1, column=1).value = "Sheet"
        alerts_sheet.cell(row=1, column=2).value = "Supplier"
        alerts_sheet.cell(row=1, column=3).value = "Invoice #"
        alerts_sheet.cell(row=1, column=4).value = "Field"
        alerts_sheet.cell(row=1, column=5).value = "Value"
        alerts_sheet.cell(row=1, column=6).value = "Alert"
        
        # Format headers
        header_fill = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
        for col in range(1, 7):
            cell = alerts_sheet.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = header_fill
        
        # Set column widths
        alerts_sheet.column_dimensions['A'].width = 20
        alerts_sheet.column_dimensions['B'].width = 30
        alerts_sheet.column_dimensions['C'].width = 15
        alerts_sheet.column_dimensions['D'].width = 15
        alerts_sheet.column_dimensions['E'].width = 10
        alerts_sheet.column_dimensions['F'].width = 40
    else:
        # No alerts, add a message
        alerts_sheet.cell(row=1, column=1).value = "No suspicious values detected in this report."
    
    # Save and return
    buffer.seek(0)
    return buffer.getvalue(), alerts