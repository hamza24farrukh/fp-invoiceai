import os
import re
import json
import tempfile
import time
from io import BytesIO
from datetime import datetime
# Note: use app.logger for logging (Flask built-in)
import pandas as pd
from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, flash, session
from werkzeug.utils import secure_filename
from dotenv import load_dotenv
from openpyxl.styles import Font, PatternFill, NamedStyle

# Local imports
from pathlib import Path
from pdf_processor import extract_text_from_pdf, extract_text_from_document
from ai_extractor import extract_data_from_document, extract_data_from_text, suggest_categories
from supplier_manager import SupplierManager
from invoice_manager import InvoiceManager
from excel_exporter import export_to_excel, export_suppliers_to_excel, export_invoice_data_to_excel, export_monthly_financial_report
from excel_import import import_suppliers_from_excel
from bank_statement_converter import BankStatementConverter
from currency_manager import (
    load_exchange_rates, update_exchange_rate,
    get_exchange_rates_for_display, convert_to_pkr,
    parse_amount, format_amount, get_amount_with_conversions
)

# Load environment variables
load_dotenv()

app = Flask(__name__)
app.secret_key = os.urandom(24)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max upload size

# Helper method to get a descriptive label for the processing type based on options
def _get_processing_type_label(options):
    """
    Generate a descriptive label for the bank statement processing type based on selected options.
    
    Args:
        options: Dictionary with processing options (advanced_extraction, ai_transaction_matching)
        
    Returns:
        String description of the processing type
    """
    if not options:
        return "Standard Processing"
        
    features = []
    
    if options.get('advanced_extraction', False):
        features.append("Advanced Payee Extraction")
        
    if options.get('use_ai_transaction_matching', False):
        features.append("AI Transaction Matching")
        
    if not features:
        return "Standard Processing"
        
    return f"Enhanced Processing ({', '.join(features)})"

# Create uploads folder if it doesn't exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Initialize supplier manager
supplier_manager = SupplierManager()

# Initialize invoice manager
invoice_manager = InvoiceManager()

INVOICES_JSON_PATH = os.path.join(os.path.dirname(__file__), 'invoices.json')


def sync_invoices_snapshot():
    """Persist current invoice state to invoices.json."""
    try:
        invoice_manager.export_invoices_to_json(
            supplier_manager=supplier_manager,
            output_path=INVOICES_JSON_PATH
        )
        app.logger.info("Invoices snapshot synced to %s", INVOICES_JSON_PATH)
    except Exception as exc:
        app.logger.error(f"Failed to sync invoices.json: {exc}")

# Helper functions
def allowed_file(filename):
    allowed_extensions = {
        'pdf', 'xlsx', 'xls', 'jpg', 'jpeg', 'png', 'bmp', 'tiff', 'tif', 'gif',
        'mp3', 'wav', 'm4a', 'ogg', 'flac', 'webm'
    }
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/model_status')
def model_status():
    """Return JSON with availability of each AI model."""
    from ai_extractor import _check_ollama_available
    from pdf_processor import get_ocr_info

    status = {
        'gemini': {
            'name': 'Gemini 2.0 Flash',
            'available': bool(os.getenv('GOOGLE_API_KEY')),
            'type': 'cloud'
        },
        'mistral': {
            'name': 'Mistral Medium',
            'available': bool(os.getenv('MISTRAL_API_KEY')),
            'type': 'cloud'
        },
        'ollama': {
            'name': 'Ollama Phi-3.5',
            'available': _check_ollama_available("phi3.5"),
            'type': 'local'
        },
        'tesseract': {
            'name': 'Tesseract OCR',
            'available': get_ocr_info().get('available', False),
            'type': 'local'
        },
        'dit': {
            'name': 'DiT Classifier',
            'available': True,
            'type': 'local'
        }
    }
    return jsonify(status)

@app.route('/upload', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        # Read processing mode from form
        processing_mode = request.form.get('processing_mode', 'auto')
        app.logger.info(f"Processing mode: {processing_mode}")

        # Check if any file was uploaded
        if 'files[]' not in request.files:
            flash('No file part')
            return redirect(request.url)

        files = request.files.getlist('files[]')
        
        # Check if any file was selected
        if not files or files[0].filename == '':
            flash('No selected file')
            return redirect(request.url)
        
        # Process each file
        all_extracted_data = []
        new_suppliers = []
        invoices_changed = False
        
        # Initialize document classifier once for all files
        from document_classifier import DocumentClassifier
        classifier = DocumentClassifier()

        # Process files sequentially with rate limiting
        for i, file in enumerate(files):
            # If this is not the first file, add a delay to prevent quota issues
            if i > 0:
                wait_time = 8  # Increased wait time to 8 seconds between files
                app.logger.info(f"Rate limiting: Waiting {wait_time} seconds before processing file {i+1}/{len(files)}")
                time.sleep(wait_time)

            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(file_path)

                try:
                    app.logger.info(f"Processing file {i+1}/{len(files)}: {filename}")

                    # Get file extension to determine processing method
                    file_ext = Path(filename).suffix.lower()

                    # --- Document Classification (DiT Vision Transformer) ---
                    doc_type_result = None
                    try:
                        doc_type_result = classifier.classify(file_path)
                        app.logger.info(
                            f"Document classified as: {doc_type_result.get('document_type')} "
                            f"(confidence: {doc_type_result.get('confidence', 0):.0%})"
                        )
                    except Exception as cls_err:
                        app.logger.warning(f"Document classification skipped: {cls_err}")
                        doc_type_result = {'document_type': 'invoice', 'confidence': 0.0}

                    # --- Tesseract OCR Pre-processing for images ---
                    ocr_text = ""
                    if file_ext in ['.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif', '.gif']:
                        try:
                            from pdf_processor import extract_text_with_ocr
                            ocr_text = extract_text_with_ocr(file_path)
                            if ocr_text:
                                app.logger.info(f"OCR pre-extracted {len(ocr_text)} chars from {filename}")
                        except Exception as ocr_err:
                            app.logger.warning(f"OCR pre-processing skipped: {ocr_err}")
                    
                    # Get all existing supplier names to help with extraction
                    existing_suppliers = []
                    all_suppliers = supplier_manager.get_all_suppliers()
                    if all_suppliers:
                        # Create a list of just the supplier names
                        existing_suppliers = [s.get('supplier_name', '') for s in all_suppliers if s.get('supplier_name')]
                        app.logger.info(f"Found {len(existing_suppliers)} existing suppliers in database")
                    
                    # Add retry logic for API quota errors
                    max_retries = 5  # Increased retries
                    retry_count = 0
                    extracted_data = None
                    
                    while retry_count < max_retries and extracted_data is None:
                        try:
                            # For PDFs and images, use direct document processing with Gemini vision
                            if file_ext in ['.pdf', '.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif', '.gif']:
                                app.logger.info(f"Processing {file_ext} file with direct document processing")
                                extracted_data = extract_data_from_document(
                                    file_path=file_path,
                                    document_title=filename,
                                    existing_suppliers=existing_suppliers,
                                    ocr_text=ocr_text,
                                    processing_mode=processing_mode
                                )
                            else:
                                # For Excel files, use text-based processing (legacy method)
                                app.logger.info(f"Processing Excel file with text-based method")
                                extracted_text = f"Excel file: {filename}\nThis is a financial statement in Excel format."
                                extracted_data = extract_data_from_text(
                                    text=extracted_text,
                                    source_type="Excel",
                                    document_title=filename,
                                    existing_suppliers=existing_suppliers,
                                    processing_mode=processing_mode
                                )
                            
                            # If no data was extracted but we haven't hit max retries, try again
                            if not extracted_data:
                                if retry_count < max_retries - 1:
                                    retry_count += 1
                                    wait_time = 8 + (retry_count * 3)  # Progressive waiting
                                    app.logger.warning(f"No data extracted, retry {retry_count}/{max_retries} for {filename} in {wait_time}s")
                                    time.sleep(wait_time)
                                else:
                                    app.logger.warning(f"Failed to extract data from {filename} after {max_retries} attempts")
                                    flash(f"No data could be extracted from {filename} after multiple attempts")
                        except Exception as e:
                            retry_count += 1
                            err_msg = str(e)
                            
                            # If hit quota limit (HTTP 429), wait longer before retry
                            if "429" in err_msg and retry_count < max_retries:
                                # Extract wait time from error message if available
                                retry_delay_match = re.search(r'retry_delay\s+{\s+seconds:\s+(\d+)', err_msg)
                                suggested_wait = int(retry_delay_match.group(1)) if retry_delay_match else None
                                
                                # Use suggested wait time or calculate our own with exponential backoff
                                wait_time = suggested_wait or min(15 * retry_count, 60)  # Increased backoff, max 60 seconds
                                app.logger.warning(f"Rate limit hit, retrying in {wait_time}s ({retry_count}/{max_retries})")
                                time.sleep(wait_time)
                            elif retry_count < max_retries:
                                app.logger.error(f"Error during extraction: {err_msg}, retry {retry_count}/{max_retries}")
                                time.sleep(10)  # Increased wait time for general errors
                            else:
                                app.logger.error(f"Max retries reached for {filename}. Last error: {err_msg}")
                                flash(f"Failed to process {filename} after {max_retries} attempts. Please try again later.")
                                # Don't raise here, just continue with the next file
                                break
                    
                    if extracted_data:
                        # Check if supplier is new and convert data to our expected format
                        for raw_data in extracted_data:
                            # Map fields from the RAG extraction format to our internal format
                            # Now focusing ONLY on the 4 essential fields as per requirements
                            supplier_data = {
                                # KEY FIELD 1: Supplier Name (Transactor)
                                "supplier_name": raw_data.get("Transactor"),
                                # KEY FIELD 2: Service Category
                                "category": raw_data.get("Expense Category Account"),
                                # KEY FIELD 3: Transaction Type
                                "transaction_type": raw_data.get("Kind of Transaction"),
                                # KEY FIELD 4: VAT Number (extract and clean)
                                "vat_number": raw_data.get("VAT") or "",
                                "tax_amount": raw_data.get("VAT") or "",  # Keep for backwards compatibility
                                # Store original extraction for reference
                                "raw_extraction": raw_data,
                                # Currency is needed for proper VAT display
                                "currency": "PKR"
                            }
                            
                            # Create a list of categories if needed
                            if supplier_data["category"]:
                                supplier_data["categories"] = [supplier_data["category"]]
                            
                            # Check if this is a new supplier
                            if supplier_data["supplier_name"]:
                                # Try to get existing supplier data
                                existing_supplier = supplier_manager.get_supplier(supplier_data["supplier_name"])
                                
                                if existing_supplier:
                                    # Save newly extracted VAT number if existing supplier doesn't have one
                                    extracted_vat = supplier_data.get("vat_number")
                                    existing_vat = existing_supplier.get("vat_number")
                                    
                                    vat_updated = False
                                    if extracted_vat and (not existing_vat or existing_vat.strip() == ""):
                                        # If we found a VAT number and the supplier doesn't have one yet, update it
                                        existing_supplier["vat_number"] = extracted_vat
                                        existing_supplier["tax_amount"] = extracted_vat  # For backward compatibility
                                        app.logger.info(f"Added VAT number '{extracted_vat}' to existing supplier '{supplier_data['supplier_name']}'")
                                        vat_updated = True
                                        # Add a flag to the supplier_data that will be used in the results page
                                        supplier_data["vat_updated"] = True
                                    
                                    # ENHANCED: Apply existing categories and transaction type to the new data
                                    # Always get the most up-to-date supplier information from the database
                                    refreshed_supplier = supplier_manager.get_supplier(supplier_data["supplier_name"])
                                    
                                    app.logger.info(f"Using updated supplier information for '{supplier_data['supplier_name']}'")
                                    
                                    # Use the latest category information from the database
                                    if refreshed_supplier and refreshed_supplier.get("categories"):
                                        supplier_data["categories"] = refreshed_supplier.get("categories")
                                        app.logger.info(f"Applied latest categories from database: {supplier_data['categories']}")
                                    elif existing_supplier.get("categories"):
                                        supplier_data["categories"] = existing_supplier.get("categories")
                                    
                                    # Also update the primary category field
                                    if refreshed_supplier and refreshed_supplier.get("category"):
                                        supplier_data["category"] = refreshed_supplier.get("category")
                                        app.logger.info(f"Applied latest category from database: {supplier_data['category']}")
                                    elif existing_supplier.get("category"):
                                        supplier_data["category"] = existing_supplier.get("category")
                                        
                                    # Use the latest transaction type from the database
                                    if refreshed_supplier and refreshed_supplier.get("transaction_type"):
                                        supplier_data["transaction_type"] = refreshed_supplier.get("transaction_type")
                                        app.logger.info(f"Applied latest transaction type from database: {supplier_data['transaction_type']}")
                                    elif existing_supplier.get("transaction_type"):
                                        supplier_data["transaction_type"] = existing_supplier.get("transaction_type")
                                    
                                    # If the VAT number was updated, keep it in the supplier_data too
                                    if vat_updated:
                                        supplier_data["vat_number"] = extracted_vat
                                        supplier_data["tax_amount"] = extracted_vat
                                    else:
                                        # Otherwise, use existing VAT number from the refreshed supplier data
                                        if refreshed_supplier and refreshed_supplier.get("vat_number"):
                                            # Prioritize the refreshed supplier data
                                            supplier_data["vat_number"] = refreshed_supplier.get("vat_number")
                                            supplier_data["tax_amount"] = refreshed_supplier.get("vat_number")
                                            app.logger.info(f"Applied latest VAT number from database: {supplier_data['vat_number']}")
                                        elif existing_vat:
                                            # Fall back to the existing supplier data if refreshed doesn't have it
                                            supplier_data["vat_number"] = existing_vat
                                            supplier_data["tax_amount"] = existing_vat
                                        
                                    # Update supplier with the new data while preserving categorization info
                                    supplier_manager.update_supplier(supplier_data["supplier_name"], supplier_data)
                                else:
                                    # This is a new supplier
                                    new_suppliers.append(supplier_data)
                                    supplier_manager.add_supplier(supplier_data)
                                
                                # Extract and store invoice details
                                try:
                                    # Parse invoice date - try to find a date in the document
                                    invoice_date = None
                                    date_str = None
                                    
                                    # First, try to get invoice date from AI extraction
                                    if 'raw_extraction' in supplier_data and 'Invoice Date' in supplier_data['raw_extraction'] and supplier_data['raw_extraction']['Invoice Date']:
                                        extracted_date = supplier_data['raw_extraction']['Invoice Date']
                                        if extracted_date and extracted_date.strip():
                                            app.logger.info(f"Found Invoice Date in AI extraction: {extracted_date}")
                                            
                                            # Check if it's already in YYYY-MM-DD format
                                            if re.match(r'^\d{4}-\d{1,2}-\d{1,2}$', extracted_date):
                                                invoice_date = extracted_date
                                                app.logger.info(f"Using AI-extracted invoice date: {invoice_date}")
                                            else:
                                                # Try to parse the extracted date in various formats
                                                for fmt in ['%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d', 
                                                          '%d-%m-%y', '%Y-%m-%d', '%m/%d/%Y', '%b %d, %Y', '%B %d, %Y']:
                                                    try:
                                                        invoice_date = datetime.strptime(extracted_date, fmt).strftime('%Y-%m-%d')
                                                        app.logger.info(f"Parsed AI-extracted date: {invoice_date} from {extracted_date}")
                                                        break
                                                    except ValueError:
                                                        continue
                                    
                                    # If AI extraction didn't work, try to get date information from the filename
                                    if not invoice_date:
                                        # Patterns to look for: "1-1-2023" or "Year/Month/Day" format
                                        date_pattern = r'(\d{1,2}[-/]\d{1,2}[-/]\d{2,4}|\d{4}[-/]\d{1,2}[-/]\d{1,2})'
                                        
                                        # Try to find a date in the filename
                                        filename_date_match = re.search(date_pattern, filename)
                                        if filename_date_match:
                                            date_str = filename_date_match.group(1)
                                        
                                        if date_str:
                                            # Try different date formats
                                            try:
                                                # Try different formats: DD-MM-YYYY, YYYY-MM-DD, etc.
                                                for fmt in ['%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d', 
                                                          '%d-%m-%y', '%y-%m-%d', '%d/%m/%y', '%y/%m/%d']:
                                                    try:
                                                        invoice_date = datetime.strptime(date_str, fmt).strftime('%Y-%m-%d')
                                                        app.logger.info(f"Parsed invoice date from filename: {invoice_date} from {date_str}")
                                                        break
                                                    except ValueError:
                                                        continue
                                            except Exception as date_err:
                                                app.logger.warning(f"Could not parse date {date_str}: {date_err}")
                                    
                                    # Special handling for Upwork invoices
                                    is_upwork_invoice = supplier_data.get('supplier_name') and 'upwork' in supplier_data.get('supplier_name', '').lower()
                                    
                                    if is_upwork_invoice:
                                        app.logger.info("Processing Upwork invoice with special handling")
                                        
                                        # Try to extract date from invoice number/transaction ID for Upwork
                                        if 'raw_extraction' in supplier_data and 'description' in supplier_data.get('raw_extraction', {}):
                                            description = supplier_data['raw_extraction']['description']
                                            app.logger.info(f"Checking Upwork description for date: {description}")
                                            
                                            # Look for date patterns in the description 
                                            date_pattern = r'(\d{1,2}[-/]\d{1,2}[-/]\d{2,4}|\d{4}[-/]\d{1,2}[-/]\d{1,2})'
                                            date_match = re.search(date_pattern, str(description))
                                            if date_match:
                                                upwork_date_str = date_match.group(1)
                                                app.logger.info(f"Found date in Upwork description: {upwork_date_str}")
                                                
                                                # Try to parse the date
                                                for fmt in ['%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d', 
                                                          '%d-%m-%y', '%y-%m-%d', '%d/%m/%y', '%y/%m/%d']:
                                                    try:
                                                        invoice_date = datetime.strptime(upwork_date_str, fmt).strftime('%Y-%m-%d')
                                                        app.logger.info(f"Parsed Upwork date: {invoice_date} from {upwork_date_str}")
                                                        break
                                                    except ValueError:
                                                        continue
                                        
                                        # If we couldn't extract a date, use a default date for Upwork invoices (2023-03-15)
                                        if not invoice_date:
                                            invoice_date = '2023-03-15'  # Use a default date for Upwork invoices
                                            app.logger.info(f"Using default date for Upwork invoice: {invoice_date}")
                                    
                                    # If no date found by any method, use today's date
                                    if not invoice_date:
                                        invoice_date = datetime.now().strftime('%Y-%m-%d')
                                        app.logger.info(f"No date found, using today's date: {invoice_date}")
                                    
                                    # Extract invoice number
                                    invoice_number = None
                                    
                                    # First, try to get invoice number from AI extraction
                                    if 'raw_extraction' in supplier_data and 'Invoice Number' in supplier_data['raw_extraction'] and supplier_data['raw_extraction']['Invoice Number']:
                                        invoice_number = supplier_data['raw_extraction']['Invoice Number']
                                        app.logger.info(f"Using AI-extracted invoice number: {invoice_number}")
                                    
                                    # If no AI extraction, look for patterns like Facture-6355 or INV-12345 in filename
                                    if not invoice_number:
                                        inv_pattern = r'(?:invoice|facture|inv|bill|order)[-\s\.:#]?(\d+)'
                                        filename_inv_match = re.search(inv_pattern, filename.lower())
                                        if filename_inv_match:
                                            invoice_number = filename_inv_match.group(1)
                                            app.logger.info(f"Extracted invoice number from filename pattern: {invoice_number}")
                                    
                                    # If still no match, try to extract any number sequence from the filename
                                    if not invoice_number:
                                        number_match = re.search(r'(\d{4,})', filename)  # Look for 4+ digit sequences
                                        if number_match:
                                            invoice_number = number_match.group(1)
                                            app.logger.info(f"Extracted numeric sequence as invoice number: {invoice_number}")
                                    
                                    # If no invoice number found, use the filename as reference
                                    if not invoice_number:
                                        # Remove extension from filename
                                        invoice_number = os.path.splitext(filename)[0]
                                        app.logger.info(f"Using filename as invoice number: {invoice_number}")
                                    
                                    # Extract financial data from the raw extraction
                                    # Try to get amount, VAT, and total values
                                    amount = 0.0
                                    vat = 0.0
                                    total_bgn = 0.0
                                    total_euro = 0.0
                                    currency = 'PKR'  # Default currency
                                    
                                    # Get values from raw extraction if available
                                    if 'raw_extraction' in supplier_data:
                                        raw_data = supplier_data['raw_extraction']
                                        
                                        # Special handling for Upwork invoices
                                        if is_upwork_invoice:
                                            app.logger.info("Processing Upwork financial data with special handling")
                                            
                                            # Check if description contains payment info for debugging
                                            if 'description' in raw_data:
                                                app.logger.info(f"Upwork description: {raw_data.get('description', '')}")
                                            
                                            # For Upwork, first try to detect the currency from the description
                                            description = str(raw_data.get('description', ''))
                                            
                                            # Try to identify currency in description
                                            if '$' in description or 'USD' in description.upper():
                                                currency = 'USD'
                                                app.logger.info("Detected USD currency in Upwork description")
                                            elif '€' in description or 'EUR' in description.upper():
                                                currency = 'EUR'
                                                app.logger.info("Detected EUR currency in Upwork description")
                                            elif 'BGN' in description.upper() or 'лв' in description:
                                                currency = 'BGN'
                                                app.logger.info("Detected BGN currency in Upwork description")
                                            elif '₨' in description or 'Rs' in description or 'PKR' in description.upper():
                                                currency = 'PKR'
                                                app.logger.info("Detected PKR currency in Upwork description")
                                                
                                            # Look for monetary amounts in the description - format like $30.00, €30.00, or just numbers
                                            amount_pattern = r'[\$€]?(\d+(?:\.\d+)?)'
                                            amount_matches = re.findall(amount_pattern, description)
                                            
                                            if amount_matches:
                                                for potential_amount in amount_matches:
                                                    try:
                                                        potential_value = float(potential_amount)
                                                        # Use amount if it seems reasonable (greater than 1)
                                                        if potential_value >= 1:
                                                            amount = potential_value
                                                            app.logger.info(f"Extracted amount {amount} from Upwork description")
                                                            break
                                                    except ValueError:
                                                        continue
                                        
                                        # Standard amount processing for all invoices
                                        if amount == 0 and 'Amount' in raw_data and raw_data['Amount']:
                                            try:
                                                # Strip any non-numeric characters and convert to float
                                                amount_str = str(raw_data['Amount']).strip()
                                                amount_str = re.sub(r'[^\d\.]', '', amount_str)
                                                if amount_str:
                                                    amount = float(amount_str)
                                            except (ValueError, TypeError):
                                                app.logger.warning(f"Could not parse Amount: {raw_data['Amount']}")
                                        
                                        # Get VAT Amount (tax) - check different possible field names
                                        tax_field_name = None
                                        if 'Tax Amount' in raw_data and raw_data['Tax Amount']:
                                            tax_field_name = 'Tax Amount'
                                        elif 'VAT Amount' in raw_data and raw_data['VAT Amount']:
                                            tax_field_name = 'VAT Amount'
                                            
                                        if tax_field_name:
                                            try:
                                                # Strip any non-numeric characters and convert to float
                                                vat_str = str(raw_data[tax_field_name]).strip()
                                                vat_str = re.sub(r'[^\d\.]', '', vat_str)
                                                if vat_str:
                                                    vat = float(vat_str)
                                                    app.logger.info(f"Extracted VAT/Tax Amount: {vat} from field {tax_field_name}")
                                            except (ValueError, TypeError):
                                                app.logger.warning(f"Could not parse {tax_field_name}: {raw_data[tax_field_name]}")
                                        
                                        # For Upwork, we usually don't have VAT
                                        if is_upwork_invoice and vat == 0:
                                            app.logger.info("Setting VAT to 0 for Upwork invoice")
                                        
                                        # Get Total
                                        total_amount = 0.0
                                        if 'Total' in raw_data and raw_data['Total']:
                                            try:
                                                # Strip any non-numeric characters and convert to float
                                                total_str = str(raw_data['Total']).strip()
                                                total_str = re.sub(r'[^\d\.]', '', total_str)
                                                if total_str:
                                                    total_amount = float(total_str)
                                            except (ValueError, TypeError):
                                                app.logger.warning(f"Could not parse Total: {raw_data['Total']}")
                                        
                                        # If we have total but missing amount or VAT, calculate the missing values
                                        if total_amount > 0:
                                            if amount > 0 and vat == 0:
                                                # If we have amount but no VAT, calculate VAT
                                                vat = total_amount - amount
                                            elif amount == 0 and vat > 0:
                                                # If we have VAT but no amount, calculate amount
                                                amount = total_amount - vat
                                            elif amount == 0 and vat == 0:
                                                # If both are missing, assume total is the amount (no VAT)
                                                amount = total_amount
                                        
                                        # Get Currency (if not already set by Upwork special handling)
                                        if not is_upwork_invoice and 'Currency' in raw_data and raw_data['Currency']:
                                            currency_str = str(raw_data['Currency']).strip().upper()
                                            if currency_str in ['BGN', 'EUR', 'USD', 'GBP', 'PKR']:
                                                currency = currency_str

                                        # Calculate totals based on currency using exchange rates from currency_manager
                                        # PKR is the base reporting currency
                                        rates = load_exchange_rates()
                                        eur_to_pkr = rates.get('EUR_TO_PKR', 280.50)
                                        usd_to_pkr = rates.get('USD_TO_PKR', 278.50)
                                        bgn_to_pkr = rates.get('BGN_TO_PKR', 143.40)

                                        total_original = amount + vat
                                        if currency == 'PKR':
                                            total_bgn = total_original  # total_bgn column stores PKR base amount
                                            total_euro = total_original  # total_euro column stores PKR base amount
                                        elif currency == 'EUR':
                                            total_euro = total_original * eur_to_pkr
                                            total_bgn = total_euro
                                        elif currency == 'USD':
                                            total_euro = total_original * usd_to_pkr
                                            total_bgn = total_euro
                                        elif currency == 'BGN':
                                            total_euro = total_original * bgn_to_pkr
                                            total_bgn = total_euro
                                        else:
                                            # For other currencies, convert to PKR
                                            total_euro, _ = convert_to_pkr(total_original, currency)
                                            total_bgn = total_euro
                                    
                                    # Extract description if available
                                    description = ""
                                    if 'raw_extraction' in supplier_data:
                                        raw_data = supplier_data['raw_extraction']
                                        
                                        # Try different potential field names for description
                                        if 'Description' in raw_data and raw_data['Description'] and str(raw_data['Description']).strip() not in ['0', 'null', 'None', 'undefined']:
                                            description = str(raw_data['Description']).strip()
                                        elif 'Service Description' in raw_data and raw_data['Service Description'] and str(raw_data['Service Description']).strip() not in ['0', 'null', 'None', 'undefined']:
                                            description = str(raw_data['Service Description']).strip()
                                        elif 'Item Description' in raw_data and raw_data['Item Description'] and str(raw_data['Item Description']).strip() not in ['0', 'null', 'None', 'undefined']:
                                            description = str(raw_data['Item Description']).strip()
                                        elif 'Product Description' in raw_data and raw_data['Product Description'] and str(raw_data['Product Description']).strip() not in ['0', 'null', 'None', 'undefined']:
                                            description = str(raw_data['Product Description']).strip()
                                        elif 'Line Items' in raw_data and raw_data['Line Items'] and str(raw_data['Line Items']).strip() not in ['0', 'null', 'None', 'undefined']:
                                            description = str(raw_data['Line Items']).strip()
                                            
                                        # If description is still empty, check if there's any information in "Expense Category Account" we can use
                                        if not description and 'Expense Category Account' in raw_data and raw_data['Expense Category Account'] and str(raw_data['Expense Category Account']).strip() not in ['0', 'null', 'None', 'undefined']:
                                            category_info = str(raw_data['Expense Category Account']).strip()
                                            # Only use if it has meaningful detail (more than just a generic category name)
                                            if len(category_info) > 15 or "," in category_info or " - " in category_info:
                                                description = f"Service: {category_info}"
                                    
                                    if description:
                                        app.logger.info(f"Extracted description: '{description}'")
                                    else:
                                        app.logger.info("No description extracted from invoice")
                                    
                                    # Create invoice data
                                    invoice_data = {
                                        'date': invoice_date,
                                        'invoice_number': invoice_number,
                                        'transactor': supplier_data.get('supplier_name'),
                                        'amount': amount,
                                        'vat': vat,
                                        'total_bgn': total_bgn,
                                        'total_euro': total_euro,
                                        'currency': currency,
                                        'notes': f"Extracted from {filename}",
                                        'description': description,  # Add the extracted description
                                        'file_path': file_path,  # Store the file path for PDF viewing
                                        'raw_extraction': supplier_data.get('raw_extraction')
                                    }
                                    
                                    # Store invoice in database
                                    invoice_id = invoice_manager.add_invoice(invoice_data)
                                    if invoice_id > 0:
                                        app.logger.info(f"Added invoice {invoice_number} for {supplier_data.get('supplier_name')} with ID {invoice_id}")
                                        # Add invoice ID to data for display
                                        supplier_data['invoice_id'] = invoice_id
                                        supplier_data['invoice_number'] = invoice_number
                                        supplier_data['invoice_date'] = invoice_date
                                        invoices_changed = True
                                    else:
                                        app.logger.warning(f"Failed to add invoice for {supplier_data.get('supplier_name')}")
                                        
                                except Exception as inv_err:
                                    app.logger.error(f"Error adding invoice: {inv_err}")
                            
                            all_extracted_data.append(supplier_data)
                    
                    # Don't remove file - we need it for PDF viewing
                    # os.remove(file_path)
                    
                except Exception as e:
                    # Log the error but continue processing other files
                    error_message = f'Error processing {filename}: {str(e)}'
                    app.logger.error(error_message)
                    flash(error_message)
                    # Continue with next file instead of stopping the entire batch
                    continue
        
        # If we processed any files successfully, show results
        if invoices_changed:
            sync_invoices_snapshot()

        if all_extracted_data:
            # Create a list to track suppliers with updated VAT numbers
            updated_vat_numbers = []
            
            # Loop through the extracted data to find updates to VAT numbers
            for data in all_extracted_data:
                # If supplier is in all_extracted_data but not in new_suppliers,
                # it's an existing supplier that may have been updated
                if data['supplier_name'] and data['vat_number'] and data['supplier_name'] not in [s['supplier_name'] for s in new_suppliers]:
                    # Check if this supplier had its VAT number updated
                    # We'll rely on our logged information from the vat_updated flag
                    if 'vat_updated' in data and data['vat_updated']:
                        updated_vat_numbers.append({
                            'supplier_name': data['supplier_name'],
                            'vat_number': data['vat_number']
                        })
            
            # Store data in session
            session['extracted_data'] = all_extracted_data
            session['new_suppliers'] = new_suppliers
            session['updated_vat_numbers'] = updated_vat_numbers
            
            return redirect(url_for('results'))
        else:
            flash('No data could be successfully extracted from any of the uploaded files')
            return redirect(request.url)
    
    return render_template('upload.html')

@app.route('/results')
def results():
    extracted_data = session.get('extracted_data', [])
    new_suppliers = session.get('new_suppliers', [])
    updated_vat_numbers = session.get('updated_vat_numbers', [])
    
    return render_template('results.html', 
                         extracted_data=extracted_data,
                         new_suppliers=new_suppliers,
                         updated_vat_numbers=updated_vat_numbers)

@app.route('/suppliers')
def suppliers():
    page = request.args.get('page', 1, type=int)
    per_page = 20  # Number of suppliers per page
    
    # Get all suppliers
    all_suppliers = supplier_manager.get_all_suppliers()
    
    # Calculate pagination
    total_suppliers = len(all_suppliers)
    total_pages = (total_suppliers + per_page - 1) // per_page  # Ceiling division
    
    # Adjust page number if it's out of range
    if page < 1:
        page = 1
    elif page > total_pages and total_pages > 0:
        page = total_pages
    
    # Slice the list to get the current page's suppliers
    start_idx = (page - 1) * per_page
    end_idx = min(start_idx + per_page, total_suppliers)
    current_page_suppliers = all_suppliers[start_idx:end_idx]
    
    return render_template('suppliers.html', 
                         suppliers=current_page_suppliers,
                         page=page,
                         total_pages=total_pages,
                         total_suppliers=total_suppliers,
                         max=max,
                         min=min)
    
@app.route('/invoices')
def invoices():
    # Get filtering parameters from request (with defaults)
    year = request.args.get('year', None)
    month = request.args.get('month', None)
    
    # Get current year and all years from invoices for the filter dropdown
    current_year = datetime.now().year
    current_month = datetime.now().month
    
    # Get all distinct years from invoices for filter dropdown
    all_invoices = invoice_manager.get_all_invoices()
    invoice_years = sorted(set(int(invoice['date'].split('-')[0]) for invoice in all_invoices 
                        if invoice.get('date') and len(invoice['date'].split('-')) >= 1), reverse=True)
    
    # If no invoice years, use current year
    if not invoice_years:
        invoice_years = [current_year]
    
    # If year and month are specified, filter invoices
    if year and month and year.isdigit() and month.isdigit():
        year_int = int(year)
        month_int = int(month)
        filtered_invoices = invoice_manager.get_expense_invoices_by_year_month(year_int, month_int)
        app.logger.info(f"Filtered expense invoices for {year_int}-{month_int}: {len(filtered_invoices)} records")
    elif year and year.isdigit():
        # Filter by year only
        year_int = int(year)
        # Get all invoices for the selected year
        all_year_invoices = invoice_manager.get_invoices_by_year(year_int)
        # Filter to only expense invoices
        filtered_invoices = [invoice for invoice in all_year_invoices 
                            if invoice.get('is_income') == 0 or invoice.get('is_income') is None]
        app.logger.info(f"Filtered expense invoices for {year_int}: {len(filtered_invoices)} records")
    else:
        # No filtering, get all expense invoices
        filtered_invoices = invoice_manager.get_expense_invoices()
    
    # For template rendering
    selected_year = int(year) if year and year.isdigit() else current_year
    selected_month = int(month) if month and month.isdigit() else None
    
    # Enhance invoices with supplier information
    for invoice in filtered_invoices:
        if invoice.get('transactor'):
            supplier = supplier_manager.get_supplier(invoice['transactor'])
            if supplier:
                # Add expense category from supplier data
                if 'category' in supplier:
                    invoice['expense_category'] = supplier['category']
                elif 'categories' in supplier and supplier['categories']:
                    # Join multiple categories with commas
                    invoice['expense_category'] = ', '.join(supplier['categories'])
                
                # Add account description (kind of transaction)
                invoice['account_description'] = supplier.get('transaction_type', '')
    
    return render_template('invoices.html', 
                         invoices=filtered_invoices, 
                         section="expenses",
                         years=invoice_years,
                         months=range(1, 13),
                         selected_year=selected_year,
                         selected_month=selected_month,
                         current_year=current_year,
                         current_month=current_month,
                         supplier_manager=supplier_manager)

@app.route('/income')
def income():
    # Get filtering parameters from request (with defaults)
    year = request.args.get('year', None)
    month = request.args.get('month', None)
    
    # Get current year and all years from invoices for the filter dropdown
    current_year = datetime.now().year
    current_month = datetime.now().month
    
    # Get all distinct years from invoices for filter dropdown
    all_invoices = invoice_manager.get_all_invoices()
    invoice_years = sorted(set(int(invoice['date'].split('-')[0]) for invoice in all_invoices 
                        if invoice.get('date') and len(invoice['date'].split('-')) >= 1), reverse=True)
    
    # If no invoice years, use current year
    if not invoice_years:
        invoice_years = [current_year]
    
    # If year and month are specified, filter invoices
    if year and month and year.isdigit() and month.isdigit():
        year_int = int(year)
        month_int = int(month)
        filtered_invoices = invoice_manager.get_income_invoices_by_year_month(year_int, month_int)
        app.logger.info(f"Filtered income invoices for {year_int}-{month_int}: {len(filtered_invoices)} records")
    elif year and year.isdigit():
        # Filter by year only
        year_int = int(year)
        # Get all invoices for the selected year
        all_year_invoices = invoice_manager.get_invoices_by_year(year_int)
        # Filter to only income invoices
        filtered_invoices = [invoice for invoice in all_year_invoices if invoice.get('is_income') == 1]
        app.logger.info(f"Filtered income invoices for {year_int}: {len(filtered_invoices)} records")
    else:
        # No filtering, get all income invoices
        filtered_invoices = invoice_manager.get_income_invoices()
    
    # For template rendering
    selected_year = int(year) if year and year.isdigit() else current_year
    selected_month = int(month) if month and month.isdigit() else None
    
    return render_template('income.html', 
                         invoices=filtered_invoices, 
                         section="income",
                         years=invoice_years,
                         months=range(1, 13),
                         selected_year=selected_year,
                         selected_month=selected_month,
                         current_year=current_year,
                         current_month=current_month)

@app.route('/invoice_details/<invoice_number>')
def invoice_details(invoice_number):
    """
    Display detailed information about a specific invoice.
    
    Args:
        invoice_number: The invoice number to display details for
    """
    invoice = invoice_manager.find_invoice_by_number(invoice_number)
    if not invoice:
        flash(f"Invoice {invoice_number} not found", "danger")
        return redirect(url_for('invoices'))
    
    # Ensure transaction_type has a value for proper display
    if not invoice.get('transaction_type'):
        if invoice.get('is_income'):
            invoice['transaction_type'] = 'INCOME'
        else:
            invoice['transaction_type'] = 'EXPENSE'
    
    # Fetch supplier details to get expense categories and account details
    supplier_info = None
    if invoice.get('transactor'):
        supplier_info = supplier_manager.get_supplier(invoice['transactor'])
        
        # Add supplier details to the invoice data for the template
        if supplier_info:
            # Get expense category from supplier (could be a single category or list of categories)
            if 'category' in supplier_info:
                invoice['expense_category'] = supplier_info['category']
            elif 'categories' in supplier_info and supplier_info['categories']:
                # Join multiple categories with commas
                invoice['expense_category'] = ', '.join(supplier_info['categories'])
            else:
                invoice['expense_category'] = 'Not specified'
    
    # Add currency conversion data for display
    # First ensure we have the right field names for backward compatibility
    if 'vat' in invoice and 'vat_amount' not in invoice:
        invoice['vat_amount'] = invoice['vat']
    
    # Make sure tax_amount is also available for EVN and other special invoices
    # This ensures the Tax Amount is visible in the VAT field on the invoice details page
    if invoice.get('vat') and 'tax_amount' not in invoice:
        invoice['tax_amount'] = invoice['vat']
    
    if 'total_amount' not in invoice:
        amount = float(invoice.get('amount', 0) or 0)
        vat = float(invoice.get('vat', 0) or invoice.get('vat_amount', 0) or 0)
        invoice['total_amount'] = amount + vat
    
    # Get account description (Kind of Transaction)
    if supplier_info:
        invoice['account_description'] = supplier_info.get('transaction_type', 'Not specified')
        
        # Additional supplier data that might be useful
        invoice['vat_number'] = supplier_info.get('vat_number', '')
    
    # Ensure description is available (from notes or separate description field)
    if not invoice.get('description') and invoice.get('notes'):
        invoice['description'] = invoice['notes']
    elif not invoice.get('description'):
        invoice['description'] = 'No description available'
    
    # Handle currency conversions for financial fields
    currency = invoice.get('currency', 'PKR')
    
    # Process amount with conversions if present
    if invoice.get('amount'):
        try:
            invoice['amount_with_conversions'] = get_amount_with_conversions(invoice['amount'], currency)
        except Exception as e:
            app.logger.error(f"Error converting amount: {str(e)}")
            
    # Process VAT amount with conversions if present
    if invoice.get('vat_amount'):
        try:
            invoice['vat_amount_with_conversions'] = get_amount_with_conversions(invoice['vat_amount'], currency)
        except Exception as e:
            app.logger.error(f"Error converting VAT amount: {str(e)}")
            
    # Process total amount with conversions if present
    if invoice.get('total_amount'):
        try:
            invoice['total_amount_with_conversions'] = get_amount_with_conversions(invoice['total_amount'], currency)
        except Exception as e:
            app.logger.error(f"Error converting total amount: {str(e)}")
    
    # Get exchange rates for display
    exchange_rates = get_exchange_rates_for_display()
        
    app.logger.info(f"Displaying details for invoice: {invoice_number}")
    return render_template('invoice_details.html', 
                          invoice=invoice, 
                          supplier_info=supplier_info,
                          exchange_rates=exchange_rates)

@app.route('/view_invoice_pdf/<invoice_id>')
def view_invoice_pdf(invoice_id):
    """
    Display the PDF file associated with an invoice.
    
    Args:
        invoice_id: The ID of the invoice to view the PDF for
    """
    app.logger.info(f"Viewing PDF for invoice ID: {invoice_id}")
    
    try:
        invoice = invoice_manager.get_invoice(int(invoice_id))
        
        if not invoice:
            flash("Invoice not found", "danger")
            return redirect(url_for('invoices'))
        
        # Check if this invoice has a PDF file
        if not invoice.get('file_path'):
            # Look for a PDF file with the invoice number
            potential_files = []
            for folder in [app.config['UPLOAD_FOLDER'], 'attached_assets']:
                for file in os.listdir(folder):
                    if file.lower().endswith('.pdf') and invoice['invoice_number'].lower() in file.lower():
                        potential_files.append(os.path.join(folder, file))
            
            if not potential_files:
                flash("No PDF file associated with this invoice", "warning")
                return redirect(url_for('invoice_details', invoice_number=invoice['invoice_number']))
            
            # Use the first matching file found
            pdf_path = potential_files[0]
            
            # Update the invoice with this file path for future use
            invoice_manager.update_invoice(invoice['id'], {'file_path': pdf_path})
            app.logger.info(f"Updated invoice {invoice_id} with PDF path: {pdf_path}")
        else:
            pdf_path = invoice['file_path']
        
        # If the file doesn't exist, look in the uploads folder
        if not os.path.exists(pdf_path) and not os.path.isabs(pdf_path):
            pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], pdf_path)
        
        # If the file still doesn't exist, check if it's in the attached_assets folder
        if not os.path.exists(pdf_path):
            filename = os.path.basename(pdf_path)
            alt_path = os.path.join('attached_assets', filename)
            if os.path.exists(alt_path):
                pdf_path = alt_path
        
        # If we still can't find the file, show error
        if not os.path.exists(pdf_path):
            flash(f"PDF file not found: {os.path.basename(pdf_path)}", "danger")
            return redirect(url_for('invoice_details', invoice_number=invoice['invoice_number']))
        
        app.logger.info(f"Serving PDF file: {pdf_path}")
        return send_file(
            pdf_path,
            mimetype='application/pdf'
        )
    except Exception as e:
        app.logger.error(f"Error displaying PDF: {str(e)}")
        flash(f"Error displaying PDF: {str(e)}", "danger")
        return redirect(url_for('invoices'))

@app.route('/edit_supplier/<supplier_name>', methods=['GET', 'POST'])
def edit_supplier(supplier_name):
    supplier = supplier_manager.get_supplier(supplier_name)
    
    if request.method == 'POST':
        updated_data = {
            'supplier_name': request.form.get('supplier_name'),
            'category': request.form.get('category'),
            'transaction_type': request.form.get('transaction_type'),
            'vat_number': request.form.get('vat_number')
        }
        
        # Handle multiple categories
        additional_categories = request.form.getlist('additional_categories[]')
        
        # Combine primary category and additional categories, removing duplicates
        all_categories = [updated_data['category']] if updated_data['category'] else []
        
        for cat in additional_categories:
            if cat and cat not in all_categories:
                all_categories.append(cat)
        
        # Store all categories
        updated_data['categories'] = all_categories
        
        # Add other fields from original supplier data
        for key, value in supplier.items():
            if key not in updated_data:
                updated_data[key] = value
        
        # Update supplier
        supplier_manager.update_supplier(supplier_name, updated_data)
        flash(f'Supplier {updated_data["supplier_name"]} updated successfully')
        return redirect(url_for('suppliers'))
    
    return render_template('edit_supplier.html', supplier=supplier)

@app.route('/delete_supplier/<supplier_name>', methods=['POST'])
def delete_supplier(supplier_name):
    # URL decode the supplier name
    import urllib.parse
    decoded_supplier_name = urllib.parse.unquote(supplier_name)
    
    success = supplier_manager.delete_supplier(decoded_supplier_name)
    
    if success:
        flash(f'Supplier {decoded_supplier_name} deleted successfully')
    else:
        flash(f'Failed to delete supplier {decoded_supplier_name}. Supplier not found.', 'error')
        
    return redirect(url_for('suppliers'))

@app.route('/delete_all_suppliers', methods=['POST'])
def delete_all_suppliers():
    supplier_manager.delete_all_suppliers()
    flash('All suppliers deleted successfully')
    return redirect(url_for('suppliers'))
    
@app.route('/delete_all_invoices', methods=['POST'])
def delete_all_invoices():
    """Delete all invoices from the database"""
    app.logger.info("Delete all invoices requested")
    if invoice_manager.delete_all_invoices():
        sync_invoices_snapshot()
        flash('All invoices deleted successfully')
    else:
        flash('Failed to delete invoices', 'error')
    return redirect(url_for('invoices'))

@app.route('/edit_invoice/<int:invoice_id>', methods=['GET', 'POST'])
def edit_invoice(invoice_id):
    # Check if we're creating a new invoice (id=0)
    if invoice_id == 0:
        # Determine if this is an income entry based on the referrer
        is_income = 'income' in request.referrer if request.referrer else False
        transaction_type = 'INCOME' if is_income else 'EXPENSES with VAT'
        
        # Create a new empty invoice object
        invoice = {
            'id': 0,
            'date': datetime.now().strftime('%Y-%m-%d'),
            'invoice_number': '',
            'transactor': '',
            'amount': 0.0,
            'vat_amount': 0.0,
            'total_amount': 0.0,
            'total_bgn': 0.0,
            'total_euro': 0.0,
            'currency': 'PKR',
            'notes': '',
            'description': '',
            'is_income': is_income,
            'transaction_type': transaction_type
        }
    else:
        # Load existing invoice
        invoice = invoice_manager.get_invoice(invoice_id)
        
        if not invoice:
            flash(f'Invoice with ID {invoice_id} not found', 'error')
            return redirect(url_for('invoices'))
            
        # Handle legacy fields - map vat to vat_amount if needed
        if 'vat' in invoice and 'vat_amount' not in invoice:
            invoice['vat_amount'] = invoice['vat']
            
        # Calculate total_amount if it doesn't exist
        if 'total_amount' not in invoice:
            amount = float(invoice.get('amount', 0) or 0)
            vat = float(invoice.get('vat', 0) or invoice.get('vat_amount', 0) or 0)
            invoice['total_amount'] = amount + vat
    
    if request.method == 'POST':
        # Check if this is an income transaction
        is_income = request.form.get('is_income') == 'on'
        
        # Handle currency conversion
        currency = request.form.get('currency', 'PKR')
        amount = request.form.get('amount') or '0'
        vat_amount = request.form.get('vat') or '0'
        total_amount = request.form.get('total_amount') or '0'

        # Parse values, handling both comma and period as decimal separators
        parsed_amount = parse_amount(amount)
        parsed_vat = parse_amount(vat_amount)
        parsed_total = parse_amount(total_amount)

        # Load exchange rates
        rates = load_exchange_rates()

        # Convert to PKR (base currency)
        pkr_amount, error_amount = convert_to_pkr(parsed_amount, currency)
        pkr_vat, error_vat = convert_to_pkr(parsed_vat, currency)
        pkr_total, error_total = convert_to_pkr(parsed_total, currency)

        # Prepare the updated data
        # total_bgn and total_euro DB columns now both store PKR base amount
        updated_data = {
            'date': request.form.get('date'),
            'invoice_number': request.form.get('invoice_number'),
            'transactor': request.form.get('transactor'),
            'amount': parsed_amount,
            'vat_amount': parsed_vat,
            'total_amount': parsed_total,
            'total_bgn': pkr_total,
            'total_euro': pkr_total,
            'currency': currency,
            'notes': request.form.get('notes'),
            'description': request.form.get('description'),
            'is_income': is_income,
            'transaction_type': request.form.get('transaction_type')
        }
        
        # For backward compatibility, also set the 'vat' field
        updated_data['vat'] = parsed_vat
        
        # Update invoice
        success = invoice_manager.update_invoice(invoice_id, updated_data)
        
        if success:
            sync_invoices_snapshot()
            flash(f'Invoice {updated_data["invoice_number"]} updated successfully')
        else:
            flash(f'Failed to update invoice {updated_data["invoice_number"]}', 'error')
        
        # Redirect to the appropriate page based on whether it's an income or expense
        if updated_data.get('is_income'):
            return redirect(url_for('income'))
        else:
            return redirect(url_for('invoices'))
    
    # Get exchange rates for the template
    exchange_rates = get_exchange_rates_for_display()
    
    return render_template('edit_invoice.html', 
                          invoice=invoice, 
                          exchange_rates=exchange_rates)

@app.route('/delete_invoice/<int:invoice_id>', methods=['GET', 'POST'])
def delete_invoice(invoice_id):
    invoice = invoice_manager.get_invoice(invoice_id)
    
    if not invoice:
        flash(f'Invoice with ID {invoice_id} not found', 'error')
        return redirect(url_for('invoices'))
    
    success = invoice_manager.delete_invoice(invoice_id)
    
    if success:
        sync_invoices_snapshot()
        flash(f'Invoice {invoice["invoice_number"]} deleted successfully')
    else:
        flash(f'Failed to delete invoice with ID {invoice_id}', 'error')
    
    # Redirect to the appropriate page based on whether it was an income or expense
    if invoice.get('is_income'):
        return redirect(url_for('income'))
    else:
        return redirect(url_for('invoices'))

@app.route('/comprehensive_report', methods=['GET', 'POST'])
def comprehensive_report():
    """
    Generate a comprehensive financial report combining invoice data and bank statement data.
    This report includes income, expenses, and bank transactions with detailed categorization.
    """
    # Get all years and months for the dropdown
    all_invoices = invoice_manager.get_all_invoices()
    
    # Extract unique years from invoices
    invoice_years = sorted(set(int(invoice['date'].split('-')[0]) for invoice in all_invoices 
                      if invoice.get('date') and len(invoice['date'].split('-')) >= 1), reverse=True)
    
    # If no invoice years, use current year
    if not invoice_years:
        invoice_years = [datetime.now().year]
    
    # Current month and year for default selection
    current_year = datetime.now().year
    current_month = datetime.now().month
    current_month_name = datetime.now().strftime("%B")
    
    # Month names for selection dropdown
    month_names = ["January", "February", "March", "April", "May", "June", 
                   "July", "August", "September", "October", "November", "December"]
    
    if request.method == 'POST':
        # Get selected month and year
        selected_month = request.form.get('month', current_month_name)
        selected_year = request.form.get('year', str(current_year))
        
        # Initialize bank_statement_df to None - will be populated if a file is uploaded
        bank_statement_df = None
        app.logger.info("Initializing bank_statement_df to None at the start of POST processing")
        
        # Convert selected month to month number for database query
        try:
            selected_month_num = month_names.index(selected_month) + 1
            selected_year_num = int(selected_year)
        except (ValueError, TypeError) as e:
            app.logger.error(f"Invalid month or year selection: {e}")
            flash(f"Invalid month or year selection: {e}", "danger")
            return redirect(url_for('comprehensive_report'))
        
        # Get invoices for the specific month and year
        app.logger.info(f"Generating comprehensive report for {selected_month} {selected_year}")
        invoices = invoice_manager.get_invoices_by_year_month(selected_year_num, selected_month_num)
        
        if not invoices:
            flash(f'No invoices found for {selected_month} {selected_year}', 'warning')
            return redirect(url_for('comprehensive_report'))
        
        # Check if a bank statement file was uploaded
        if 'bank_statement_file' in request.files:
            bank_file = request.files['bank_statement_file']
            if bank_file and bank_file.filename and '.' in bank_file.filename:
                try:
                    # Save the file temporarily
                    filename = secure_filename(bank_file.filename)
                    bank_file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                    bank_file.save(bank_file_path)
                    
                    # Load the bank statement data
                    try:
                        # Use BankStatementConverter to auto-detect format and load
                        try:
                            converter = BankStatementConverter()
                            format_info = converter.detect_format(bank_file_path)
                            bank_statement_df = converter.convert(bank_file_path)
                            app.logger.info(f"Loaded bank statement with {len(bank_statement_df)} rows (format: {format_info.get('detected_format', 'Unknown')})")
                        except Exception as conv_err:
                            app.logger.warning(f"BankStatementConverter failed, falling back to pandas: {conv_err}")
                            bank_statement_df = pd.read_excel(bank_file_path)
                            app.logger.info(f"Loaded bank statement with {len(bank_statement_df)} rows (direct load)")

                        if 'Payee' not in bank_statement_df.columns:
                            app.logger.info("Bank statement loaded but Payee column not found - using raw data")
                        else:
                            app.logger.info("Using pre-processed bank statement (Payee column found)")
                            
                    except Exception as e:
                        app.logger.error(f"Error loading bank statement: {str(e)}")
                        flash(f"Error loading bank statement: {str(e)}", "danger")
                except Exception as e:
                    app.logger.error(f"Error saving bank statement file: {str(e)}")
                    flash(f"Error with bank statement file: {str(e)}", "danger")
        
        # Create the comprehensive report
        from excel_exporter import create_comprehensive_financial_report
        
        try:
            # Debug log about bank_statement_df
            if bank_statement_df is not None:
                app.logger.info(f"Bank statement DataFrame exists with {len(bank_statement_df)} rows and columns: {list(bank_statement_df.columns)}")
            else:
                app.logger.warning("Bank statement DataFrame is None - no bank data will be included in report")
            
            # Pass supplier_manager to ensure we have up-to-date supplier information
            excel_file, alerts = create_comprehensive_financial_report(
                invoices, 
                bank_statement_df=bank_statement_df,
                month=selected_month,
                year=selected_year,
                supplier_manager=supplier_manager  # Pass supplier_manager for current categories
            )
            
            # Create a temporary file to serve
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            temp_file.write(excel_file)
            temp_file.close()
            
            # Display alerts to the user if any suspicious values were found
            if alerts:
                app.logger.warning(f"Found {len(alerts)} suspicious values in the report")
                for alert in alerts:
                    flash(alert, "warning")
            
            # Build a descriptive file name
            file_name = f"financial_report_{selected_year}_{selected_month}.xlsx"
            
            return send_file(
                temp_file.name,
                as_attachment=True,
                download_name=file_name,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        except Exception as e:
            app.logger.error(f"Error creating comprehensive report: {str(e)}")
            flash(f"Error creating report: {str(e)}", "danger")
            return redirect(url_for('comprehensive_report'))
    
    return render_template('comprehensive_report.html',
                         years=invoice_years,
                         months=month_names,
                         current_year=current_year,
                         current_month=current_month_name)

@app.route('/export', methods=['GET', 'POST'])
def export():
    if request.method == 'POST':
        export_type = request.form.get('export_type')
        
        if export_type == 'suppliers':
            suppliers = supplier_manager.get_all_suppliers()
            if not suppliers:
                flash('No suppliers available to export')
                return redirect(request.url)
            
            excel_file, alerts = export_suppliers_to_excel(suppliers)
            
            # Create a temporary file to serve
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            temp_file.write(excel_file)
            temp_file.close()
            
            # Display alerts to the user if any suspicious values were found
            if alerts:
                app.logger.warning(f"Found {len(alerts)} suspicious values in the supplier export")
                for alert in alerts:
                    flash(alert, "warning")
            
            return send_file(
                temp_file.name,
                as_attachment=True,
                download_name='suppliers.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
        elif export_type == 'invoices':
            # Export the full invoice database
            invoices = invoice_manager.get_all_invoices()
            if not invoices:
                flash('No invoices available to export')
                return redirect(request.url)
            
            # Convert to pandas DataFrame
            df = pd.DataFrame(invoices)
            
            # Create Excel file in memory
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Invoices', index=False)
                
                # Get the workbook and sheet
                workbook = writer.book
                worksheet = writer.sheets['Invoices']
                
                # Set column widths
                for idx, col in enumerate(df.columns):
                    max_len = max(df[col].astype(str).map(len).max(), len(col)) + 3
                    worksheet.column_dimensions[chr(65 + idx)].width = min(max_len, 30)
                
                # Add some formatting
                header_font = Font(bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                
                # Apply header formatting
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Format date columns
                date_format = NamedStyle(name='date_style', number_format='YYYY-MM-DD')
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=df.columns.get_loc('date') + 1)
                    cell.style = date_format
                
                # Format currency columns
                currency_format = NamedStyle(name='currency_style', number_format='#,##0.00')
                currency_columns = ['amount', 'vat', 'total_bgn', 'total_euro']
                for col_name in currency_columns:
                    if col_name in df.columns:
                        col_idx = df.columns.get_loc(col_name) + 1
                        for row in range(2, len(df) + 2):
                            cell = worksheet.cell(row=row, column=col_idx)
                            cell.style = currency_format
            
            # Get the Excel file from the output
            output.seek(0)
            excel_data = output.getvalue()
            
            # Create a temporary file to serve
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            temp_file.write(excel_data)
            temp_file.close()
            
            return send_file(
                temp_file.name,
                as_attachment=True,
                download_name='invoice_database.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
        elif export_type == 'extracted_data':
            extracted_data = session.get('extracted_data', [])
            if not extracted_data:
                flash('No extracted data available to export')
                return redirect(request.url)
            
            excel_file = export_invoice_data_to_excel(extracted_data)
            
            # Create a temporary file to serve
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            temp_file.write(excel_file)
            temp_file.close()
            
            return send_file(
                temp_file.name,
                as_attachment=True,
                download_name='invoice_data.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
        elif export_type == 'monthly_report':
            # Get all suppliers data for the report
            all_suppliers = supplier_manager.get_all_suppliers()
            
            # Also get any extracted data from the current session
            session_data = session.get('extracted_data', [])
            
            # Combine all data - give priority to session data as it's more recent
            combined_data = all_suppliers.copy()
            
            # Add session data if not already in the list (by invoice number)
            existing_invoice_numbers = [s.get('invoice_number') for s in combined_data if s.get('invoice_number')]
            for item in session_data:
                if item.get('invoice_number') not in existing_invoice_numbers:
                    combined_data.append(item)
            
            if not combined_data:
                flash('No data available to generate monthly report')
                return redirect(request.url)
            
            # Generate the monthly report
            current_date = datetime.now()
            month_name = current_date.strftime("%B")
            year = current_date.year
            
            excel_file = export_monthly_financial_report(combined_data)
            
            # Create a temporary file to serve
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            temp_file.write(excel_file)
            temp_file.close()
            
            return send_file(
                temp_file.name,
                as_attachment=True,
                download_name=f'INCOME-EXPENSES {month_name} {year}.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
    
    return render_template('export.html')

@app.route('/api/suggest_categories/<supplier_name>')
def api_suggest_categories(supplier_name):
    categories = suggest_categories(supplier_name)
    return jsonify(categories)

@app.route('/import_excel', methods=['GET', 'POST'])
def import_excel():
    """
    Handle Excel file upload for supplier information extraction.
    This route allows users to upload Excel files containing supplier information,
    which will be automatically extracted and stored in the supplier database.
    """
    if request.method == 'POST':
        # Check if a file was uploaded
        if 'excel_file' not in request.files:
            flash('No file part')
            return redirect(request.url)
            
        file = request.files['excel_file']
        
        # Check if any file was selected
        if file.filename == '':
            flash('No selected file')
            return redirect(request.url)
            
        # Check if it's an allowed file type
        if file and file.filename.lower().endswith(('.xlsx', '.xls')):
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            
            try:
                # Save the file
                file.save(file_path)
                app.logger.info(f"Processing Excel file: {filename}")
                
                # Import suppliers from the Excel file
                result = import_suppliers_from_excel(file_path, supplier_manager)
                
                # Clean up file
                os.remove(file_path)
                
                # Check for errors
                if result.get('error'):
                    flash(f"Error processing {filename}: {result['error']}")
                    return redirect(request.url)
                    
                # Success message
                flash(f"Successfully processed {filename}. "
                      f"Added {result['new_suppliers']} new suppliers and "
                      f"updated {result['updated_suppliers']} existing suppliers.")
                      
                # Redirect to the suppliers page to see the results
                return redirect(url_for('suppliers'))
                
            except Exception as e:
                error_message = f"Error processing {filename}: {str(e)}"
                app.logger.error(error_message)
                flash(error_message)
                return redirect(request.url)
        else:
            flash('Invalid file type. Please upload an Excel file (.xlsx or .xls)')
            return redirect(request.url)
            
    # GET request: Show the upload form
    return render_template('import_excel.html')

@app.route('/process_bank_statements', methods=['GET', 'POST'])
def process_bank_statements():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected')
            return redirect(request.url)

        file = request.files['file']
        if file.filename == '':
            flash('No file selected')
            return redirect(request.url)

        if file and file.filename.endswith(('.xlsx', '.xls', '.csv')):
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)

            try:
                converter = BankStatementConverter()

                # Auto-detect format
                format_info = converter.detect_format(file_path)

                # Convert to standard format
                statement_df = converter.convert(file_path)

                # Get invoices for matching
                invoice_mgr = InvoiceManager()
                all_invoices = invoice_mgr.get_all_invoices()

                # Match with invoices
                matched_df = converter.match_with_invoices(statement_df, all_invoices)

                # Store results in session
                session['bank_statement_stats'] = {
                    'total_transactions': len(statement_df),
                    'matched_transactions': len(matched_df[matched_df['matched_supplier'].notna()]) if 'matched_supplier' in matched_df.columns else 0,
                    'bank_format': format_info.get('detected_format', 'Unknown')
                }
                session['processed_file'] = filename

                # Export processed result
                output_filename = f"processed_{filename}"
                output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
                matched_df.to_excel(output_path, index=False)
                session['processed_file'] = output_filename

                flash(f'Successfully processed {len(statement_df)} transactions')
                return redirect(url_for('bank_statement_results'))

            except Exception as e:
                app.logger.error(f"Error processing bank statement: {str(e)}")
                flash(f'Error processing file: {str(e)}')
                return redirect(request.url)
        else:
            flash('Please upload an Excel (.xlsx, .xls) or CSV file')
            return redirect(request.url)

    return render_template('process_bank_statements.html')

@app.route('/bank_statement_results')
def bank_statement_results():
    """
    Show the results of bank statement processing.
    
    Supports both standard processing and enhanced AI payee extraction.
    """
    # Get processing stats from session
    stats = session.get('bank_statement_stats', {})
    
    if not stats:
        flash('No bank statement processing results available')
        return redirect(url_for('process_bank_statements'))
    
    # Check if processed file exists (new format)
    processed_file = session.get('processed_file')
    if processed_file:
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], processed_file)
        if not os.path.exists(file_path):
            flash('Processed bank statement file not found')
            return redirect(url_for('process_bank_statements'))
    else:
        # Fallback to legacy format
        output_path = stats.get('output_path', '')
        if not os.path.exists(output_path):
            flash('Processed bank statement file not found')
            return redirect(url_for('process_bank_statements'))
    
    return render_template('bank_statement_results.html', stats=stats)

@app.route('/download_processed_statement/<filename>')
def download_processed_statement(filename):
    """
    Download the processed bank statement file.
    """
    app.logger.info(f"Download request for {filename}")
    app.logger.info(f"Session data: processed_file={session.get('processed_file')}, bank_statement_stats={session.get('bank_statement_stats')}")
    
    # Try all possible places to find the file path
    output_path = None
    
    # First try direct processed_file path (preferred)
    if 'processed_file' in session:
        test_path = os.path.join(app.config['UPLOAD_FOLDER'], session['processed_file'])
        if os.path.exists(test_path):
            output_path = test_path
            app.logger.info(f"Found file using session processed_file: {output_path}")
    
    # Then try stats output_path (from older code)
    if not output_path:
        stats = session.get('bank_statement_stats', {})
        if stats and 'output_path' in stats and os.path.exists(stats['output_path']):
            output_path = stats['output_path']
            app.logger.info(f"Found file using stats output_path: {output_path}")
    
    # Then try stats processed_file path (fallback)
    if not output_path:
        stats = session.get('bank_statement_stats', {})
        if stats and 'processed_file' in stats:
            test_path = os.path.join(app.config['UPLOAD_FOLDER'], stats['processed_file'])
            if os.path.exists(test_path):
                output_path = test_path
                app.logger.info(f"Found file using stats processed_file: {output_path}")
    
    # Final check and notification
    if not output_path or not os.path.exists(output_path):
        app.logger.error(f"No file found for download: {filename}")
        flash('Processed bank statement file not found')
        return redirect(url_for('process_bank_statements'))
    
    return send_file(
        output_path,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/currency_settings', methods=['GET'])
def currency_settings():
    """
    Display currency settings page with conversion rates.
    Allows admin to update exchange rates for currency conversions.
    """
    rates = get_exchange_rates_for_display()
    return render_template('currency_settings.html', rates=rates)

@app.route('/update_exchange_rate', methods=['POST'])
def update_currency_rate():
    """
    Update a specific exchange rate.
    """
    rate_key = request.form.get('rate_key')
    rate_value = request.form.get('rate_value')
    
    if not rate_key or not rate_value or rate_key not in ['EUR_TO_PKR', 'USD_TO_PKR', 'BGN_TO_PKR']:
        flash('Invalid exchange rate information', 'error')
        return redirect(url_for('currency_settings'))
        
    try:
        # Parse and validate the rate value
        rate_value = float(rate_value.replace(',', '.'))
        if rate_value <= 0:
            flash('Exchange rate must be greater than zero', 'error')
            return redirect(url_for('currency_settings'))
            
        # Update the rate
        if update_exchange_rate(rate_key, rate_value):
            flash(f'Exchange rate updated successfully', 'success')
        else:
            flash('Failed to update exchange rate', 'error')
            
    except ValueError:
        flash('Invalid exchange rate value - must be a valid number', 'error')
        
    return redirect(url_for('currency_settings'))

# ===== NEW MODEL INTEGRATION ROUTES =====

@app.route('/upload_voice_note/<int:invoice_id>', methods=['POST'])
def upload_voice_note(invoice_id):
    """Transcribe an audio file and attach it as a note to an existing invoice."""
    invoice_manager = InvoiceManager()
    invoice = invoice_manager.get_invoice(invoice_id)

    if not invoice:
        flash('Invoice not found', 'error')
        return redirect(url_for('invoices'))

    if 'audio_file' not in request.files:
        flash('No audio file selected', 'error')
        return redirect(url_for('edit_invoice', invoice_id=invoice_id))

    audio_file = request.files['audio_file']
    if audio_file.filename == '':
        flash('No audio file selected', 'error')
        return redirect(url_for('edit_invoice', invoice_id=invoice_id))

    try:
        from audio_processor import AudioProcessor

        # Save audio file
        filename = secure_filename(audio_file.filename)
        voice_notes_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'voice_notes')
        os.makedirs(voice_notes_dir, exist_ok=True)
        audio_path = os.path.join(voice_notes_dir, f"{invoice_id}_{filename}")
        audio_file.save(audio_path)

        # Transcribe
        processor = AudioProcessor(model_size="base")
        result = processor.transcribe(audio_path)

        transcription = result.get('text', '')
        language = result.get('language', 'unknown')
        duration = result.get('duration', 0)

        # Append to invoice notes
        current_notes = invoice.get('notes', '') or ''
        separator = '\n\n' if current_notes else ''
        updated_notes = f"{current_notes}{separator}[Voice Note - {language}, {duration:.0f}s]: {transcription}"

        invoice_manager.update_invoice(invoice_id, {'notes': updated_notes})

        flash(f'Voice note transcribed successfully ({len(transcription)} characters, language: {language})', 'success')

    except Exception as e:
        app.logger.error(f"Voice note processing failed: {str(e)}")
        flash(f'Error processing voice note: {str(e)}', 'error')

    return redirect(url_for('edit_invoice', invoice_id=invoice_id))


@app.route('/upload_voice_memo', methods=['POST'])
def upload_voice_memo():
    """Transcribe a voice memo and create a new invoice entry from it."""
    if 'audio_file' not in request.files:
        flash('No audio file selected', 'error')
        return redirect(url_for('invoices'))

    audio_file = request.files['audio_file']
    if audio_file.filename == '':
        flash('No audio file selected', 'error')
        return redirect(url_for('invoices'))

    try:
        from audio_processor import AudioProcessor

        # Save audio file
        filename = secure_filename(audio_file.filename)
        voice_notes_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'voice_notes')
        os.makedirs(voice_notes_dir, exist_ok=True)
        audio_path = os.path.join(voice_notes_dir, filename)
        audio_file.save(audio_path)

        # Transcribe
        processor = AudioProcessor(model_size="base")
        result = processor.transcribe(audio_path)
        transcription = result.get('text', '')

        if not transcription:
            flash('Could not transcribe audio - no speech detected', 'warning')
            return redirect(url_for('invoices'))

        # Try to extract structured data from transcription using AI
        try:
            extracted = extract_data_from_text(transcription)
        except Exception:
            extracted = None

        # Create new invoice with transcription
        invoice_manager = InvoiceManager()
        invoice_data = {
            'date': datetime.now().strftime('%Y-%m-%d'),
            'invoice_number': '',
            'transactor': '',
            'amount': 0.0,
            'vat': 0.0,
            'total_bgn': 0.0,
            'total_euro': 0.0,
            'currency': 'PKR',
            'notes': f'[Voice Memo]: {transcription}',
            'description': transcription[:500] if transcription else '',
            'is_income': False,
            'transaction_type': 'EXPENSES with VAT'
        }

        # Merge AI-extracted data if available
        if extracted and isinstance(extracted, list) and len(extracted) > 0:
            ai_data = extracted[0] if isinstance(extracted[0], dict) else {}
            for key in ['transactor', 'invoice_number', 'date', 'amount', 'vat', 'currency']:
                if ai_data.get(key):
                    invoice_data[key] = ai_data[key]

        invoice_id = invoice_manager.add_invoice(invoice_data)

        if invoice_id > 0:
            flash(f'Voice memo transcribed and invoice created. Please review and complete the details.', 'success')
            return redirect(url_for('edit_invoice', invoice_id=invoice_id))
        else:
            flash('Failed to create invoice from voice memo', 'error')

    except Exception as e:
        app.logger.error(f"Voice memo processing failed: {str(e)}")
        flash(f'Error processing voice memo: {str(e)}', 'error')

    return redirect(url_for('invoices'))


@app.route('/model_evaluation', methods=['GET', 'POST'])
def model_evaluation():
    """Run model benchmarks and display comparison results."""
    results = None

    if request.method == 'POST':
        try:
            from model_evaluation import ModelEvaluator

            evaluator = ModelEvaluator()
            evaluator.load_ground_truth()

            # Evaluate extraction models
            try:
                def gemini_extract(file_path):
                    data = extract_data_from_document(file_path)
                    if isinstance(data, list) and len(data) > 0:
                        return data[0] if isinstance(data[0], dict) else {}
                    return data if isinstance(data, dict) else {}

                evaluator.evaluate_extraction_model(
                    model_name="Gemini 2.0 Flash (Vision + Text)",
                    extract_fn=gemini_extract,
                )
            except Exception as e:
                app.logger.warning(f"Gemini evaluation skipped: {e}")

            # Evaluate OCR
            try:
                from pdf_processor import extract_text_with_ocr

                def ocr_extract(file_path):
                    text = extract_text_with_ocr(file_path)
                    return {'raw_text': text, 'char_count': len(text)}

                evaluator.evaluate_extraction_model(
                    model_name="Tesseract OCR (Image Domain)",
                    extract_fn=ocr_extract,
                )
            except Exception as e:
                app.logger.warning(f"OCR evaluation skipped: {e}")

            # Evaluate Document Classifier
            try:
                from document_classifier import DocumentClassifier
                classifier = DocumentClassifier()
                gt = evaluator.load_ground_truth()

                class_test_data = [
                    {'file': entry['file'], 'expected_type': entry.get('expected_type', 'invoice')}
                    for entry in gt if 'expected_type' in entry
                ]

                if class_test_data:
                    evaluator.evaluate_classification_model(
                        model_name="DiT Document Classifier (Vision Transformer)",
                        classify_fn=classifier.classify,
                        test_data=class_test_data,
                    )
            except Exception as e:
                app.logger.warning(f"Classifier evaluation skipped: {e}")

            # Get comparison
            results = evaluator.compare_models()
            results['detailed'] = evaluator.results

            # Save results
            evaluator.export_results()

            flash('Model evaluation completed successfully', 'success')

        except Exception as e:
            app.logger.error(f"Model evaluation failed: {str(e)}")
            flash(f'Error during evaluation: {str(e)}', 'error')

    # Try to load previous results if no new evaluation
    if results is None:
        results_path = os.path.join('benchmarks', 'results.json')
        if os.path.exists(results_path):
            try:
                with open(results_path, 'r') as f:
                    results = json.load(f)
            except Exception:
                pass

    return render_template('model_evaluation.html', results=results)


if __name__ == '__main__':
    os.makedirs('uploads', exist_ok=True)
    os.makedirs('uploads/voice_notes', exist_ok=True)
    app.run(host='0.0.0.0', port=5000, debug=True)